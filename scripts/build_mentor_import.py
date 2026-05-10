import hashlib
import json
import re
import ssl
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\github\livelihood-ecosystem-directory")
WORKBOOK_PATH = Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\Mentors.xlsx")
OUTPUT_SQL_PATH = PROJECT_ROOT / "supabase" / "migrations" / "20260510183000_import_mentors.sql"
OUTPUT_REPORT_PATH = PROJECT_ROOT / "scripts" / "mentor_import_report.json"
CONFIG_PATH = PROJECT_ROOT / "config.js"
PUBLIC_VIEW = "ecosystem_directory_entities"
USER_AGENT = "Livelihood Ecosystem Directory Mentor Import/1.0"
SSL_CONTEXT = ssl._create_unverified_context()

SOURCE_LABEL = "Mentors workbook"
CREATED_BY_NAME = "Mentor import"
CREATED_BY_EMAIL = "tanmay@greenruraleconomy.in"
PUSA_SOURCE_URL = "https://pusakrishi.in/mentors/"

FREE_EMAIL_DOMAINS = {
    "gmail.com",
    "yahoo.com",
    "outlook.com",
    "hotmail.com",
    "rediffmail.com",
    "icloud.com",
    "ymail.com",
    "proton.me",
    "protonmail.com",
}

MANUAL_ROW_OVERRIDES = {
    "mr-rajesh-ranjan": {
        "email": "via Pusa Krishi portal",
        "phone": "via Pusa Krishi portal",
        "domain_expertise": "Entrepreneurship, rural-startup handholding, investment readiness",
        "experience": "CEO, NABVENTURES Fund; mentor for agtech, foodtech, fintech, and rural businesses",
        "business_domain": "Agtech, foodtech, fintech, rural businesses investor",
        "geography_served": "Rural India via IARI-linked networks",
        "source": "Pusa Krishi Startup Mentors",
        "website_url": PUSA_SOURCE_URL,
    }
}

PUSA_PUBLIC_PROFILES = {
    "dr-neeru-bhoosan": {
        "summary": "Pusa Krishi mentor and agricultural science expert supporting agri-tech, crop protection, and extension-oriented startups.",
        "website_url": PUSA_SOURCE_URL,
    },
    "dr-k-v-prabhu": {
        "summary": "Chairperson, Protection of Plant Varieties and Farmers’ Rights Authority, mentoring agri-mechanization and post-harvest innovation through Pusa Krishi.",
        "website_url": PUSA_SOURCE_URL,
    },
    "dr-r-n-sahoo": {
        "summary": "Principal Scientist associated with ICAR-IARI mentoring farm management, remote sensing, and rural-extension startups through Pusa Krishi.",
        "website_url": PUSA_SOURCE_URL,
    },
    "mr-rajesh-ranjan": {
        "summary": "CEO, NABVENTURES Fund, mentoring agtech, foodtech, fintech, and rural-business founders through Pusa Krishi.",
        "website_url": PUSA_SOURCE_URL,
    },
}

GENERIC_INDIA_LOCATIONS = {
    "india",
    "pan india",
    "pan-india",
    "north india",
    "remote pan india",
    "urban + rural outreach",
    "urban and rural outreach",
    "rural outreach",
}


def compact_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_unicode_text(value: str) -> str:
    return (
        str(value or "")
        .replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2011", "-")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u00a0", " ")
        .replace("\u200b", "")
    )


def clean_text(value):
    text = compact_spaces(normalize_unicode_text(value))
    if not text or text.lower() in {"na", "n/a", "none", "null", "-"}:
        return None
    return text


def slugify(value: str) -> str:
    value = re.sub(r"[^\w\s-]", "", normalize_unicode_text(value or "").strip().lower())
    value = re.sub(r"[-\s]+", "-", value)
    return value.strip("-")


def first_sentence(text: str, limit: int = 280):
    text = clean_text(text)
    if not text:
        return None
    pieces = re.split(r"(?<=[.!?])\s+", text)
    sentence = pieces[0].strip() if pieces else text.strip()
    if len(sentence) > limit:
        sentence = sentence[: limit - 1].rstrip() + "..."
    return sentence


def split_list(value):
    text = clean_text(value)
    if not text:
        return []
    return dedupe(re.split(r"\s*[,;/|]\s*", text))


def dedupe(values):
    seen = set()
    output = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        output.append(text)
    return output


def sql_text(value):
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_text_array(values):
    cleaned = [str(item).replace("'", "''") for item in values if clean_text(item)]
    if not cleaned:
        return "'{}'::text[]"
    return "ARRAY[" + ", ".join(f"'{item}'" for item in cleaned) + "]::text[]"


def sql_json(value):
    return sql_text(json.dumps(value, ensure_ascii=False)) + "::jsonb"


def read_config():
    text = CONFIG_PATH.read_text(encoding="utf-8")
    url_match = re.search(r"SUPABASE_URL:\s*'([^']+)'", text)
    key_match = re.search(r"SUPABASE_ANON_KEY:\s*'([^']+)'", text)
    if not url_match or not key_match:
        raise RuntimeError("Could not read Supabase config from config.js")
    return url_match.group(1), key_match.group(1)


def fetch_existing_mentors(base_url, anon_key):
    query = ",".join([
        "entity_uid",
        "entity_name",
        "summary",
        "description",
        "location_label",
        "primary_address",
        "district",
        "state",
        "country",
        "contact_email",
        "contact_phone",
        "website_url",
        "social_media",
        "office_locations",
        "tags",
        "keywords",
        "latitude",
        "longitude",
        "source_label",
        "source_url",
        "type_specific_data",
        "created_by_name",
        "created_by_email",
        "admin_notes",
        "search_text",
    ])
    url = (
        f"{base_url}/rest/v1/{PUBLIC_VIEW}"
        f"?select={urllib.parse.quote(query)}"
        f"&entity_type_slug=eq.mentor&limit=1000"
    )
    headers = {
        "apikey": anon_key,
        "Authorization": f"Bearer {anon_key}",
        "Accept": "application/json",
    }
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
        return json.loads(response.read().decode("utf-8"))


def load_rows():
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        raw = {
            "name": clean_text(row[0]),
            "email": clean_text(row[1]),
            "phone": clean_text(row[2]),
            "domain_expertise": clean_text(row[3]),
            "experience": clean_text(row[4]),
            "business_domain": clean_text(row[5]),
            "geography_served": clean_text(row[6]),
            "source": clean_text(row[7]),
        }
        if not raw["name"]:
            continue
        override = MANUAL_ROW_OVERRIDES.get(slugify(raw["name"]))
        if override:
            raw.update(override)
        rows.append(raw)
    return rows


def professional_website_from_email(email):
    email = clean_text(email)
    if not email or "via pusa krishi portal" in email.lower() or "@" not in email:
        return None
    domain = email.split("@", 1)[1].lower()
    if domain in FREE_EMAIL_DOMAINS:
        return None
    return f"https://{domain}"


def fetch_homepage_title(url):
    if not url:
        return None, None
    try:
        request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(request, timeout=25, context=SSL_CONTEXT) as response:
            final_url = response.geturl()
            body = response.read(300000).decode("utf-8", errors="ignore")
    except Exception:
        return url, None
    title_match = re.search(r"<title>(.*?)</title>", body, flags=re.I | re.S)
    title = compact_spaces(title_match.group(1)) if title_match else None
    return final_url, title


def infer_target_stage(text):
    haystack = (text or "").lower()
    stages = []
    if "early-stage" in haystack or "early stage" in haystack:
        stages.append("Early revenue")
    if "startup" in haystack:
        stages.append("Idea stage")
        stages.append("Prototype")
    if "sme" in haystack or "msme" in haystack:
        stages.append("Growth")
    return dedupe(stages)


def infer_mentoring_modes(email, phone):
    modes = []
    if clean_text(email) and "via pusa" not in clean_text(email).lower():
        modes.append("Email")
    if clean_text(phone) and "via pusa" not in clean_text(phone).lower():
        modes.append("Phone")
        modes.append("WhatsApp")
    if modes:
        modes.append("Video call")
    return dedupe(modes)


def normalize_geography_list(value):
    return dedupe(
        item.replace("-", " ").replace("Pan India", "Pan-India").replace("UP", "Uttar Pradesh")
        for item in split_list(value)
    )


def parse_state_and_location(geography_items):
    text = ", ".join(geography_items)
    state_map = {
        "uttar pradesh": "Uttar Pradesh",
        "odisha": "Odisha",
        "karnataka": "Karnataka",
        "rajasthan": "Rajasthan",
        "madhya pradesh": "Madhya Pradesh",
        "north india": "North India",
        "rural india": "India",
        "pan-india": "India",
    }
    state = None
    for key, label in state_map.items():
        if key in text.lower():
            state = label
            break
    location_label = ", ".join(geography_items) if geography_items else state or "India"
    if location_label == "India" and state and state != "India":
        location_label = f"{state}, India"
    return state, location_label


def geocode_query(query, cache):
    query = clean_text(query)
    if not query:
        return None
    if query in cache:
        return cache[query]
    url = f"https://nominatim.openstreetmap.org/search?format=jsonv2&limit=1&q={urllib.parse.quote(query)}"
    request = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception:
        cache[query] = None
        return None
    if not data:
        cache[query] = None
        return None
    match = data[0]
    try:
        point = (float(match["lat"]), float(match["lon"]))
    except (TypeError, KeyError, ValueError):
        cache[query] = None
        return None
    cache[query] = point
    return point


def geocode_location(location_label, state, cache):
    normalized_location = (clean_text(location_label) or "").lower()
    normalized_state = (clean_text(state) or "").lower()
    if normalized_location in GENERIC_INDIA_LOCATIONS:
        return (22.3511148, 78.6677428)
    if normalized_state == "india":
        return (22.3511148, 78.6677428)
    queries = dedupe([
        location_label,
        f"{state}, India" if state and state != "India" else state,
        "India",
    ])
    for query in queries:
        normalized_query = (clean_text(query) or "").lower()
        if normalized_query in GENERIC_INDIA_LOCATIONS:
            continue
        point = geocode_query(query, cache)
        if point:
            lat, lng = point
            if not (6 <= lat <= 38 and 68 <= lng <= 98):
                continue
            if query != "India":
                time.sleep(0.8)
            return point
    return (22.3511148, 78.6677428)


def merge_arrays(*arrays):
    merged = []
    for array in arrays:
        if isinstance(array, list):
            merged.extend(array)
    return dedupe(merged)


def build_search_text(record):
    values = [
        record.get("entity_name"),
        record.get("summary"),
        record.get("description"),
        record.get("location_label"),
        record.get("primary_address"),
        record.get("district"),
        record.get("state"),
        record.get("country"),
        " ".join(record.get("tags") or []),
        " ".join(record.get("keywords") or []),
    ]
    for value in (record.get("type_specific_data") or {}).values():
        if isinstance(value, list):
            values.append(" ".join(str(item) for item in value if clean_text(item)))
        else:
            values.append(clean_text(value) or "")
    return compact_spaces(" ".join(item for item in values if clean_text(item)))


def build_entity_uid(name):
    digest = hashlib.md5(name.encode("utf-8")).hexdigest()[:8]
    return f"mentor-{slugify(name)}-{digest}"


def build_record(row, existing, geocode_cache):
    geography_items = normalize_geography_list(row.get("geography_served"))
    state, location_label = parse_state_and_location(geography_items)
    website_url = row.get("website_url") or professional_website_from_email(row.get("email"))
    final_website_url, homepage_title = fetch_homepage_title(website_url)
    override_profile = PUSA_PUBLIC_PROFILES.get(slugify(row["name"]), {})
    summary = override_profile.get("summary") or first_sentence(
        f"{row['name']} mentors on {row.get('domain_expertise') or 'startup support'} and works across {row.get('business_domain') or row.get('geography_served') or 'entrepreneurship'}."
    )
    description = clean_text(row.get("experience")) or summary
    if homepage_title and not override_profile.get("summary"):
        description = clean_text(f"{description}. Associated organisation website: {homepage_title}") or description
    domain_expertise = split_list(row.get("domain_expertise"))
    industry_experience = split_list(row.get("business_domain"))
    email = None if row.get("email") and "via pusa" in row.get("email").lower() else row.get("email")
    phone = None if row.get("phone") and "via pusa" in row.get("phone").lower() else row.get("phone")
    type_specific = {
        "domain_expertise": domain_expertise,
        "industry_experience": industry_experience,
        "years_experience": None,
        "languages_spoken": [],
        "mentoring_modes": infer_mentoring_modes(email, phone),
        "geography_served": geography_items,
        "target_stage": infer_target_stage(" ".join([row.get("experience") or "", row.get("business_domain") or ""])),
        "availability_notes": None,
    }
    lat, lng = geocode_location(location_label, state, geocode_cache)
    source_label = row.get("source") or SOURCE_LABEL
    source_url = PUSA_SOURCE_URL if "pusa" in source_label.lower() else final_website_url
    entity_uid = (existing or {}).get("entity_uid") or build_entity_uid(row["name"])
    record = {
        "entity_uid": entity_uid,
        "entity_name": row["name"],
        "summary": summary,
        "description": description,
        "location_label": location_label,
        "primary_address": location_label,
        "district": None,
        "state": state if state and state != "North India" else None,
        "country": "India",
        "contact_email": email,
        "contact_phone": phone,
        "website_url": override_profile.get("website_url") or final_website_url,
        "social_media": {},
        "office_locations": [],
        "tags": merge_arrays(domain_expertise, industry_experience, geography_items, [source_label]),
        "keywords": merge_arrays(domain_expertise, industry_experience, geography_items, [row["name"]]),
        "latitude": lat,
        "longitude": lng,
        "source_label": SOURCE_LABEL,
        "source_url": source_url,
        "type_specific_data": type_specific,
        "created_by_name": CREATED_BY_NAME,
        "created_by_email": CREATED_BY_EMAIL,
        "admin_notes": f"Imported from Mentors.xlsx | Original source: {source_label}",
        "approval_status": "approved",
        "approved_at": "now()",
        "approved_by": "admin",
        "is_deleted": False,
        "updated_at": "now()",
        "search_text": None,
        "merge_action": "merged" if existing else "inserted",
    }
    record["search_text"] = build_search_text(record)
    return record


def build_values_sql(record):
    return "(" + ", ".join([
        sql_text(record["entity_uid"]),
        sql_text(record["entity_name"]),
        sql_text(record["summary"]),
        sql_text(record["description"]),
        sql_text(record["location_label"]),
        sql_text(record["primary_address"]),
        sql_text(record["district"]),
        sql_text(record["state"]),
        sql_text(record["country"]),
        sql_text(record["contact_email"]),
        sql_text(record["contact_phone"]),
        sql_text(record["website_url"]),
        sql_json(record["social_media"]),
        sql_json(record["office_locations"]),
        sql_text_array(record["tags"]),
        sql_text_array(record["keywords"]),
        str(record["latitude"]) if record["latitude"] is not None else "null",
        str(record["longitude"]) if record["longitude"] is not None else "null",
        sql_text(record["source_label"]),
        sql_text(record["source_url"]),
        sql_json(record["type_specific_data"]),
        sql_text(record["created_by_name"]),
        sql_text(record["created_by_email"]),
        sql_text(record["admin_notes"]),
        "'approved'",
        "now()",
        "'admin'",
        "false",
        sql_text(record["search_text"]),
        "now()",
    ]) + ")"


def build_insert_sql(records):
    return """insert into public.mentor_entities (
  entity_uid,
  entity_name,
  summary,
  description,
  location_label,
  primary_address,
  district,
  state,
  country,
  contact_email,
  contact_phone,
  website_url,
  social_media,
  office_locations,
  tags,
  keywords,
  latitude,
  longitude,
  source_label,
  source_url,
  type_specific_data,
  created_by_name,
  created_by_email,
  admin_notes,
  approval_status,
  approved_at,
  approved_by,
  is_deleted,
  search_text,
  updated_at
) values
""" + ",\n".join(build_values_sql(record) for record in records) + """
on conflict (entity_uid) do update set
  entity_name = excluded.entity_name,
  summary = excluded.summary,
  description = excluded.description,
  location_label = excluded.location_label,
  primary_address = excluded.primary_address,
  district = excluded.district,
  state = excluded.state,
  country = excluded.country,
  contact_email = excluded.contact_email,
  contact_phone = excluded.contact_phone,
  website_url = excluded.website_url,
  social_media = excluded.social_media,
  office_locations = excluded.office_locations,
  tags = excluded.tags,
  keywords = excluded.keywords,
  latitude = excluded.latitude,
  longitude = excluded.longitude,
  source_label = excluded.source_label,
  source_url = excluded.source_url,
  type_specific_data = excluded.type_specific_data,
  created_by_name = excluded.created_by_name,
  created_by_email = excluded.created_by_email,
  admin_notes = excluded.admin_notes,
  approval_status = 'approved',
  approved_at = now(),
  approved_by = 'admin',
  is_deleted = false,
  search_text = excluded.search_text,
  updated_at = now();
"""


def main():
    base_url, anon_key = read_config()
    existing_rows = fetch_existing_mentors(base_url, anon_key)
    existing_by_name = {slugify(item.get("entity_name")): item for item in existing_rows}
    geocode_cache = {}
    records = []
    report = {
        "source_label": SOURCE_LABEL,
        "row_count": 0,
        "inserted": 0,
        "merged": 0,
        "with_professional_website": 0,
        "portal_rows": 0,
    }
    for row in load_rows():
        existing = existing_by_name.get(slugify(row["name"]))
        record = build_record(row, existing, geocode_cache)
        records.append(record)
        report["row_count"] += 1
        report[record["merge_action"]] += 1
        if clean_text(record.get("website_url")):
            report["with_professional_website"] += 1
        if row.get("email") and "via pusa" in row.get("email").lower():
            report["portal_rows"] += 1
        print(f"{record['entity_name']} -> {record['merge_action']}")

    OUTPUT_SQL_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL_PATH.write_text(build_insert_sql(records) + "\n", encoding="utf-8")
    OUTPUT_REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote SQL migration to {OUTPUT_SQL_PATH}")
    print(f"Wrote import report to {OUTPUT_REPORT_PATH}")
    print(json.dumps(report, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
