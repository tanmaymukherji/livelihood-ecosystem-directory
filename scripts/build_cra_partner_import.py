import hashlib
import json
import re
import ssl
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from html import unescape
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\github\livelihood-ecosystem-directory")
WORKBOOK_PATH = Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\CRA Partner Map.xlsx")
OUTPUT_SQL_PATH = PROJECT_ROOT / "supabase" / "migrations" / "20260510123000_import_cra_partner_organisations.sql"
OUTPUT_REPORT_PATH = PROJECT_ROOT / "scripts" / "cra_partner_import_report.json"
CONFIG_PATH = PROJECT_ROOT / "config.js"
PUBLIC_VIEW = "ecosystem_directory_entities"

SOURCE_LABEL = "CRA Partner Map workbook"
CREATED_BY_NAME = "CRA partner import"
CREATED_BY_EMAIL = "tanmay@greenruraleconomy.in"

USER_AGENT = "Livelihood Ecosystem Directory CRA Import/1.0"
SSL_CONTEXT = ssl._create_unverified_context()

STATE_NAMES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh", "Goa", "Gujarat",
    "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka", "Kerala", "Madhya Pradesh",
    "Maharashtra", "Manipur", "Meghalaya", "Mizoram", "Nagaland", "Odisha", "Punjab",
    "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura", "Uttar Pradesh",
    "Uttarakhand", "West Bengal", "Delhi", "Jammu and Kashmir", "Ladakh", "Puducherry",
    "Chandigarh", "Dadra and Nagar Haveli and Daman and Diu", "Andaman and Nicobar Islands",
    "Lakshadweep",
]

REGIONAL_PRESENCE_MAP = {
    "pan": ["Pan-India", "India"],
    "south": ["South India"],
    "north": ["North India"],
    "east": ["East India"],
    "west": ["West India"],
    "central": ["Central India"],
    "northeast": ["North-East India"],
}

ORG_TYPE_MAP = {
    "grassroot/implementation csos": "cso",
    "think tanks": "cso",
    "networks/alliances": "cso",
    "funders": "csr_philanthropy",
    "media/citizen engagement": "story_teller",
}

TARGET_TABLES = {
    "cso": "cso_entities",
    "csr_philanthropy": "csr_philanthropy_entities",
    "story_teller": "story_teller_entities",
}

MANUAL_OVERRIDES = {
    "sewa-federation": {
        "entity_type_slug": "cso",
        "primary_address": "21/22, Goyal Tower, Near Jhanvi Restaurant, University Road, Panjara Pole, Ahmedabad, Gujarat 380015, India",
        "location_label": "Ahmedabad, Gujarat",
        "district": "Ahmedabad",
        "state": "Gujarat",
        "country": "India",
        "latitude": 23.0389021,
        "longitude": 72.5519454,
    },
    "vikas-centre-for-development": {
        "entity_type_slug": "cso",
        "primary_address": "ISHAVASYAM, Opp. Lajpat Nagar Society, Eeshita Tower Road, Navrangpura, Ahmedabad, Gujarat 380014, India",
        "location_label": "Ahmedabad, Gujarat",
        "district": "Ahmedabad",
        "state": "Gujarat",
        "country": "India",
        "latitude": 23.0359998,
        "longitude": 72.5643429,
    },
    "environmental-defence-fund": {
        "entity_type_slug": "cso",
        "primary_address": "257 Park Avenue South, New York, New York 10010, United States",
        "location_label": "New York, New York",
        "district": "New York County",
        "state": "New York",
        "country": "United States",
        "latitude": 40.7386142,
        "longitude": -73.9872457,
    },
}

SOCIAL_DOMAINS = {
    "youtube": "youtube.com",
    "instagram": "instagram.com",
    "linkedin": "linkedin.com",
    "facebook": "facebook.com",
    "twitter": "twitter.com",
    "x": "x.com",
}

STORY_LINK_KEYWORDS = (
    "blog",
    "story",
    "stories",
    "article",
    "articles",
    "media",
    "news",
    "video",
    "podcast",
    "publication",
)

CONTACT_PAGE_HINTS = (
    "contact",
    "reach",
    "office",
    "location",
    "about",
    "who-we-are",
    "team",
    "our-work",
    "what-we-do",
    "program",
    "programme",
    "media",
)

EMAIL_RE = re.compile(r"[\w.\-+%]+@[\w.\-]+\.[A-Za-z]{2,}", re.I)
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{7,}\d)")
URL_RE = re.compile(r"https?://[^\s\"'<>]+", re.I)
PINCODE_RE = re.compile(r"\b\d{6}\b")
FCRA_RE = re.compile(r"\bFCRA\b", re.I)
REGISTRATION_RE = re.compile(r"\b(section\s*8|society|trust|nonprofit|non-profit|ngo)\b", re.I)
GRANT_RE = re.compile(r"\b(grant|grantmaking|grant-making|funder|funding|donor|philanthrop)\b", re.I)


@dataclass
class WorkbookRow:
    organization_name: str
    about: str
    regional_presence: str
    thematic_areas: list[str]
    org_type: str
    website: str


class LinkParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.links = []
        self.title = ""
        self.meta_description = ""
        self._in_title = False
        self._skip_depth = 0
        self._text_parts = []

    def handle_starttag(self, tag, attrs):
        attrs_dict = {key.lower(): value for key, value in attrs}
        if tag.lower() in {"script", "style", "noscript"}:
            self._skip_depth += 1
        if tag.lower() == "a":
            self.links.append({
                "href": attrs_dict.get("href") or "",
                "text": "",
            })
        if tag.lower() == "meta":
            name = (attrs_dict.get("name") or attrs_dict.get("property") or "").lower()
            if name in {"description", "og:description", "twitter:description"} and not self.meta_description:
                self.meta_description = compact_spaces(unescape(attrs_dict.get("content") or ""))
        if tag.lower() == "title":
            self._in_title = True
        if tag.lower() in {"p", "div", "section", "br", "li", "h1", "h2", "h3", "h4"}:
            self._text_parts.append("\n")

    def handle_endtag(self, tag):
        if tag.lower() in {"script", "style", "noscript"} and self._skip_depth:
            self._skip_depth -= 1
        if tag.lower() == "title":
            self._in_title = False
        if tag.lower() == "a" and self.links:
            self.links[-1]["text"] = compact_spaces(self.links[-1]["text"])
        if tag.lower() in {"p", "div", "section", "li"}:
            self._text_parts.append("\n")

    def handle_data(self, data):
        if self._skip_depth:
            return
        text = compact_spaces(unescape(data or ""))
        if not text:
            return
        if self._in_title:
            self.title = compact_spaces(f"{self.title} {text}")
        if self.links and not self.links[-1]["text"]:
            self.links[-1]["text"] = text
        self._text_parts.append(text)

    @property
    def text(self):
        return compact_spaces("\n".join(self._text_parts))


def compact_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_unicode_text(value: str) -> str:
    return (
        str(value or "")
        .replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u00a0", " ")
        .replace("\u200b", "")
    )


def clean_text(value):
    text = compact_spaces(normalize_unicode_text(value))
    if not text:
        return None
    if text.lower() in {"na", "n/a", "nil", "none", "null", "-"}:
        return None
    return text


def slugify(value: str) -> str:
    value = re.sub(r"[^\w\s-]", "", normalize_unicode_text(value or "").strip().lower())
    value = re.sub(r"[-\s]+", "-", value)
    return value.strip("-")


def first_sentence(text: str, limit: int = 320):
    if not text:
        return None
    pieces = re.split(r"(?<=[.!?])\s+", text)
    sentence = pieces[0].strip() if pieces else text.strip()
    if len(sentence) > limit:
        sentence = sentence[: limit - 1].rstrip() + "..."
    return sentence


def split_list(value):
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"\s*[,;/|]\s*", text)
    return dedupe(parts)


def dedupe(values):
    seen = set()
    output = []
    for value in values:
        text = compact_spaces(value)
        key = text.lower()
        if not text or key in seen:
            continue
        seen.add(key)
        output.append(text)
    return output


def read_config():
    text = CONFIG_PATH.read_text(encoding="utf-8")
    url_match = re.search(r"SUPABASE_URL:\s*'([^']+)'", text)
    key_match = re.search(r"SUPABASE_ANON_KEY:\s*'([^']+)'", text)
    if not url_match or not key_match:
        raise RuntimeError("Could not read Supabase config from config.js")
    return url_match.group(1), key_match.group(1)


def fetch_json(url, headers=None, timeout=60):
    request = urllib.request.Request(url)
    for key, value in (headers or {}).items():
        request.add_header(key, value)
    with urllib.request.urlopen(request, timeout=timeout, context=SSL_CONTEXT) as response:
        return json.loads(response.read().decode("utf-8"))


def fetch_existing_entities(base_url, anon_key):
    query = ",".join([
        "entity_uid",
        "entity_type_slug",
        "entity_name",
        "summary",
        "description",
        "location_label",
        "primary_address",
        "district",
        "state",
        "country",
        "contact_email",
        "contact_phone",
        "website_url",
        "social_media",
        "office_locations",
        "tags",
        "keywords",
        "latitude",
        "longitude",
        "source_label",
        "source_url",
        "type_specific_data",
        "created_by_name",
        "created_by_email",
        "admin_notes",
        "search_text",
    ])
    type_filter = "in.(cso,csr_philanthropy,story_teller)"
    url = (
        f"{base_url}/rest/v1/{PUBLIC_VIEW}"
        f"?select={urllib.parse.quote(query)}"
        f"&entity_type_slug={urllib.parse.quote(type_filter)}"
        f"&limit=2000"
    )
    headers = {
        "apikey": anon_key,
        "Authorization": f"Bearer {anon_key}",
        "Accept": "application/json",
    }
    return fetch_json(url, headers=headers)


def load_workbook_rows():
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        if not any(raw):
            continue
        org_name = clean_text(raw[0])
        if not org_name:
            continue
        rows.append(WorkbookRow(
            organization_name=org_name,
            about=clean_text(raw[1]) or "",
            regional_presence=clean_text(raw[2]) or "",
            thematic_areas=split_list(raw[3]),
            org_type=clean_text(raw[4]) or "",
            website=normalize_website(raw[5]),
        ))
    return rows


def normalize_website(value):
    text = clean_text(value)
    if not text:
        return None
    if not re.match(r"^https?://", text, re.I):
        text = f"https://{text.lstrip('/')}"
    return text.rstrip("/")


def domain_key(url):
    if not url:
        return ""
    try:
        parsed = urllib.parse.urlparse(url)
    except Exception:
        return ""
    host = (parsed.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def classify_row(row: WorkbookRow):
    org_type_key = row.org_type.lower()
    if org_type_key in ORG_TYPE_MAP:
        return ORG_TYPE_MAP[org_type_key]
    if org_type_key == "international orgnaizations":
        combined = " ".join([row.organization_name, row.about, " ".join(row.thematic_areas)]).lower()
        return "csr_philanthropy" if re.search(r"\b(fund|foundation|grant|philanthrop|donor)\b", combined, re.I) else "cso"
    return "cso"


def parse_html(html):
    parser = LinkParser()
    parser.feed(html)
    return parser


def fetch_html(url, timeout=45):
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=timeout, context=SSL_CONTEXT) as response:
        final_url = response.geturl()
        body = response.read()
        charset = response.headers.get_content_charset() or "utf-8"
        return final_url, body.decode(charset, errors="ignore")


def extract_json_ld(html):
    blocks = re.findall(
        r"<script[^>]+type=[\"']application/ld\+json[\"'][^>]*>(.*?)</script>",
        html,
        flags=re.I | re.S,
    )
    items = []
    for block in blocks:
        text = compact_spaces(unescape(block))
        if not text:
            continue
        try:
            parsed = json.loads(text)
        except Exception:
            continue
        if isinstance(parsed, list):
            items.extend(parsed)
        else:
            items.append(parsed)
    return items


def html_to_text_blocks(html):
    if not html:
        return ""
    text = re.sub(r"(?is)<script.*?>.*?</script>", " ", html)
    text = re.sub(r"(?is)<style.*?>.*?</style>", " ", text)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</(p|div|section|li|ul|ol|tr|td|h1|h2|h3|h4|h5|footer|header|address)>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = unescape(text)
    text = text.replace("\xa0", " ")
    lines = [compact_spaces(line) for line in text.splitlines()]
    return "\n".join(line for line in lines if line)


def extract_social_links(links):
    socials = {}
    for item in links:
        href = normalize_website(item.get("href"))
        if not href:
            continue
        host = domain_key(href)
        for key, domain in SOCIAL_DOMAINS.items():
            if domain in host and key not in socials:
                socials[key] = href
    return socials


def extract_emails(text):
    items = []
    for match in EMAIL_RE.findall(text or ""):
        value = match.strip(".,;:()[]{}<>").lower()
        if value.endswith((".png", ".jpg", ".jpeg", ".webp")):
            continue
        items.append(value)
    return dedupe(items)


def extract_phones(text):
    output = []
    for match in PHONE_RE.findall(text or ""):
        cleaned = re.sub(r"[^\d+]", "", match)
        digits = re.sub(r"\D", "", cleaned)
        if len(digits) < 10 or len(digits) > 14:
            continue
        if cleaned.startswith("00"):
            cleaned = "+" + cleaned[2:]
        output.append(match.strip())
    return dedupe(output)


def extract_jsonld_addresses(items):
    output = []
    for item in items:
        if not isinstance(item, dict):
            continue
        for candidate in flatten_jsonld_items(item):
            address = candidate.get("address")
            if not isinstance(address, dict):
                continue
            parts = [
                clean_text(address.get("streetAddress")),
                clean_text(address.get("addressLocality")),
                clean_text(address.get("addressRegion")),
                clean_text(address.get("postalCode")),
                clean_text(address.get("addressCountry") or "India"),
            ]
            formatted = ", ".join(part for part in parts if part)
            if formatted:
                output.append(formatted)
    return dedupe(output)


def flatten_jsonld_items(item):
    output = [item]
    graph = item.get("@graph")
    if isinstance(graph, list):
        output.extend([entry for entry in graph if isinstance(entry, dict)])
    return output


def extract_address_candidates(text):
    if not text:
        return []
    raw = normalize_unicode_text(text).replace("\\n", "\n")
    lines = [compact_spaces(segment) for segment in re.split(r"[\n\r]+", raw) if compact_spaces(segment)]
    candidates = []
    for line in lines:
        lowered = line.lower()
        score = 0
        if any(keyword in lowered for keyword in ("address", "office", "registered office", "head office", "contact us")):
            score += 2
        if PINCODE_RE.search(line):
            score += 2
        if any(state.lower() in lowered for state in STATE_NAMES):
            score += 2
        if "india" in lowered:
            score += 1
        if len(line) > 30 and len(line) < 220 and len(line.split()) <= 35 and score >= 3:
            cleaned = re.sub(r"^(address|registered office|head office|office|contact us)\s*[:\-]\s*", "", line, flags=re.I)
            candidates.append(cleaned.strip(" -"))
    return dedupe(candidates)


def extract_location_mentions(text):
    if not text:
        return []
    patterns = [
        r"\bheadquartered in ([A-Z][A-Za-z.\- ]+?,\s*[A-Z][A-Za-z.\- ]+)",
        r"\bbased in ([A-Z][A-Za-z.\- ]+?,\s*[A-Z][A-Za-z.\- ]+)",
        r"\blocated in ([A-Z][A-Za-z.\- ]+?,\s*[A-Z][A-Za-z.\- ]+)",
        r"\bhead office(?: is)? in ([A-Z][A-Za-z.\- ]+?,\s*[A-Z][A-Za-z.\- ]+)",
    ]
    matches = []
    for pattern in patterns:
        for value in re.findall(pattern, text, flags=re.I):
            cleaned = compact_spaces(value).strip(" .,:;-")
            if 4 <= len(cleaned) <= 120:
                matches.append(cleaned)
    return dedupe(matches)


def pick_main_address(candidates):
    for candidate in candidates:
        cleaned = clean_text(candidate)
        if not cleaned:
            continue
        if len(cleaned) > 180:
            continue
        lowered = cleaned.lower()
        if PINCODE_RE.search(cleaned) or any(state.lower() in lowered for state in STATE_NAMES) or "india" in lowered:
            return cleaned
    for candidate in candidates:
        cleaned = clean_text(candidate)
        if cleaned and len(cleaned) <= 120:
            return cleaned
    return None


def same_domain(base_url, href):
    if not href:
        return False
    href = normalize_website(href)
    if not href:
        return False
    return domain_key(base_url) == domain_key(href)


def build_candidate_pages(base_url, links):
    pages = []
    for item in links:
        href = normalize_website(item.get("href"))
        label = compact_spaces(item.get("text"))
        if not href or not same_domain(base_url, href):
            continue
        haystack = f"{href} {label}".lower()
        if any(hint in haystack for hint in CONTACT_PAGE_HINTS):
            pages.append(href)
    fallback_paths = [
        "/contact",
        "/contact-us",
        "/about",
        "/about-us",
        "/team",
        "/media",
    ]
    base = normalize_website(base_url)
    if base:
        pages.extend([base + path for path in fallback_paths])
    return dedupe(pages)[:5]


def pick_longer(existing, incoming, minimum_gain=24):
    existing = clean_text(existing)
    incoming = clean_text(incoming)
    if not existing:
        return incoming
    if not incoming:
        return existing
    if len(incoming) >= len(existing) + minimum_gain:
        return incoming
    return existing


def merge_arrays(*arrays):
    merged = []
    for array in arrays:
        if not array:
            continue
        merged.extend([item for item in array if clean_text(item)])
    return dedupe(merged)


def merge_socials(*items):
    output = {}
    for item in items:
        if isinstance(item, dict):
            for key, value in item.items():
                if clean_text(value) and key not in output:
                    output[key] = clean_text(value)
    return output


def infer_registration_status(text):
    haystack = (text or "").lower()
    if "section 8" in haystack:
        return "Section 8"
    if "society" in haystack:
        return "Society"
    if "trust" in haystack:
        return "Trust"
    return None


def infer_target_groups(text):
    haystack = (text or "").lower()
    mapping = {
        "women": "Women",
        "youth": "Youth",
        "farmer": "Farmers",
        "tribal": "Tribal communities",
        "adolescent": "Adolescents",
        "child": "Children",
        "entrepreneur": "Entrepreneurs",
        "community": "Communities",
        "rural": "Rural communities",
    }
    return [label for key, label in mapping.items() if key in haystack]


def infer_support_instruments(text, org_type):
    values = []
    haystack = (text or "").lower()
    if org_type.lower() == "funders":
        values.append("Philanthropic grant")
    if "csr" in haystack:
        values.append("CSR grant")
    if "technical assistance" in haystack or "capacity building" in haystack:
        values.append("Technical assistance")
    if "capacity building" in haystack:
        values.append("Capacity building")
    if "volunteer" in haystack:
        values.append("Employee volunteering")
    return dedupe(values)


def infer_storytelling_modes(website_text, socials, links):
    modes = []
    if any(item for item in socials.values()):
        modes.append("Social Media")
    if "youtube" in socials:
        modes.append("Video")
    if any(any(keyword in link.lower() for keyword in ("video", "youtube", "reel")) for link in links):
        modes.append("Video")
    if any(keyword in (website_text or "").lower() for keyword in ("podcast", "audio")):
        modes.append("Audio")
    if any(keyword in (website_text or "").lower() for keyword in ("article", "blog", "story", "stories", "media", "news")) or not modes:
        modes.append("Written")
    return dedupe(modes)


def extract_known_work_links(base_url, links):
    output = []
    for link in links:
        href = normalize_website(link.get("href"))
        label = compact_spaces(link.get("text"))
        if not href or not same_domain(base_url, href):
            continue
        haystack = f"{href} {label}".lower()
        if any(keyword in haystack for keyword in STORY_LINK_KEYWORDS):
            output.append(href)
    return dedupe(output)[:8]


def extract_region_tags(value):
    return REGIONAL_PRESENCE_MAP.get((value or "").strip().lower(), [])


def normalize_query(value):
    return (
        compact_spaces(value)
        .replace("|", ", ")
        .replace(";", ", ")
        .replace(" / ", ", ")
    )


def has_usable_coordinate(latitude, longitude):
    try:
        lat = float(latitude)
        lng = float(longitude)
    except (TypeError, ValueError):
        return False
    return abs(lat) > 0.0001 or abs(lng) > 0.0001


def geocode_query(query, hints, cache):
    query = normalize_query(query)
    if not query:
        return None
    if query in cache:
        return cache[query]
    url = f"https://nominatim.openstreetmap.org/search?format=jsonv2&limit=1&q={urllib.parse.quote(query)}"
    request = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception:
        cache[query] = None
        return None
    match = data[0] if isinstance(data, list) and data else None
    if not match:
        cache[query] = None
        return None
    display_name = compact_spaces(match.get("display_name")).lower()
    if hints and not any(hint and hint in display_name for hint in hints):
        cache[query] = None
        return None
    try:
        point = (float(match["lat"]), float(match["lon"]))
    except (KeyError, TypeError, ValueError):
        cache[query] = None
        return None
    cache[query] = point if has_usable_coordinate(point[0], point[1]) else None
    return cache[query]


def geocode_entity(name, address, location_label, state, region_tags, cache):
    scoped_hints = [compact_spaces(item).lower() for item in [state] if compact_spaces(item)]
    query_options = [
        (", ".join(part for part in [address, "India"] if compact_spaces(part)), scoped_hints),
        (", ".join(part for part in [location_label, state, "India"] if compact_spaces(part)), scoped_hints),
        (", ".join(part for part in [name, state, "India"] if compact_spaces(part)), []),
        (", ".join(part for part in [name, "India"] if compact_spaces(part)), []),
        (", ".join(part for part in region_tags + ["India"] if compact_spaces(part)), scoped_hints),
    ]
    for query, hints in query_options:
        point = geocode_query(query, hints, cache)
        if point:
            return point
        time.sleep(1.0)
    return (None, None)


def extract_state(address: str, geography: list[str]):
    haystacks = [address or "", " ".join(geography or [])]
    for state in sorted(STATE_NAMES, key=len, reverse=True):
        for haystack in haystacks:
            if re.search(rf"\b{re.escape(state)}\b", haystack, flags=re.I):
                return state
    return None


def extract_district(address: str, state: str):
    if not address:
        return None
    segments = [compact_spaces(item) for item in re.split(r"[,/]", address) if compact_spaces(item)]
    if state:
        for index, segment in enumerate(segments):
            if state.lower() in segment.lower():
                if index > 0:
                    prior = re.sub(r"\b\d{4,6}\b", "", segments[index - 1]).strip(" -")
                    return compact_spaces(prior) or None
                return None
    if len(segments) >= 2:
        candidate = re.sub(r"\b\d{4,6}\b", "", segments[-2]).strip(" -")
        return compact_spaces(candidate) or None
    return None


def build_search_text(record):
    parts = [
        record.get("entity_name"),
        record.get("summary"),
        record.get("description"),
        record.get("location_label"),
        record.get("primary_address"),
        record.get("district"),
        record.get("state"),
        record.get("country"),
        " ".join(record.get("tags") or []),
        " ".join(record.get("keywords") or []),
    ]
    type_specific = record.get("type_specific_data") or {}
    for value in type_specific.values():
        if isinstance(value, list):
            parts.append(" ".join(str(item) for item in value if clean_text(item)))
        else:
            parts.append(clean_text(value) or "")
    return compact_spaces(" ".join(part for part in parts if clean_text(part)))


def sql_text(value):
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_text_array(values):
    cleaned = [str(item).replace("'", "''") for item in values if clean_text(item)]
    if not cleaned:
        return "'{}'::text[]"
    return "ARRAY[" + ", ".join(f"'{item}'" for item in cleaned) + "]::text[]"


def sql_json(value):
    return sql_text(json.dumps(value, ensure_ascii=False)) + "::jsonb"


def build_entity_uid(type_slug, name, website):
    digest = hashlib.md5((f"{name}|{website or ''}").encode("utf-8")).hexdigest()[:8]
    return f"{type_slug}-{slugify(name)}-{digest}"


def choose_existing_match(row, desired_type, existing_entities):
    normalized_name = slugify(row.organization_name)
    desired_domain = domain_key(row.website)
    exact_name_matches = []
    domain_matches = []
    for entity in existing_entities:
        entity_name_key = slugify(entity.get("entity_name"))
        entity_domain = domain_key(entity.get("website_url"))
        if desired_domain and entity_domain and desired_domain == entity_domain:
            domain_matches.append(entity)
        if normalized_name and entity_name_key == normalized_name:
            exact_name_matches.append(entity)
    for entity in domain_matches:
        if entity.get("entity_type_slug") == desired_type:
            return entity
    for entity in exact_name_matches:
        if entity.get("entity_type_slug") == desired_type:
            return entity
    return domain_matches[0] if domain_matches else (exact_name_matches[0] if exact_name_matches else None)


def enrich_from_website(row: WorkbookRow):
    if not row.website:
        return {
            "final_url": None,
            "summary": None,
            "description": None,
            "emails": [],
            "phones": [],
            "socials": {},
            "known_work_links": [],
            "addresses": [],
            "application_link": None,
            "text": "",
            "visited_pages": [],
            "notes": ["No website URL available"],
        }

    notes = []
    visited_pages = []
    combined_text_parts = []
    all_links = []
    emails = []
    phones = []
    socials = {}
    addresses = []
    summary = None
    description = None
    application_link = None

    try:
        final_url, html = fetch_html(row.website)
        visited_pages.append(final_url)
        parser = parse_html(html)
        html_text = html_to_text_blocks(html)
        jsonld = extract_json_ld(html)
        combined_text_parts.extend([parser.text, html_text])
        all_links.extend(parser.links)
        emails.extend(extract_emails(html + "\n" + parser.text + "\n" + html_text))
        phones.extend(extract_phones(html + "\n" + parser.text + "\n" + html_text))
        socials = merge_socials(socials, extract_social_links(parser.links))
        addresses = merge_arrays(
            addresses,
            extract_jsonld_addresses(jsonld),
            extract_address_candidates(html_text),
            extract_location_mentions(html_text),
        )
        summary = pick_longer(summary, parser.meta_description, minimum_gain=0)
        description = pick_longer(description, parser.text[:1200], minimum_gain=80)
        for candidate_page in build_candidate_pages(final_url, parser.links):
            if candidate_page in visited_pages:
                continue
            try:
                _, page_html = fetch_html(candidate_page)
            except Exception:
                continue
            visited_pages.append(candidate_page)
            page_parser = parse_html(page_html)
            page_text = html_to_text_blocks(page_html)
            page_jsonld = extract_json_ld(page_html)
            combined_text_parts.extend([page_parser.text, page_text])
            all_links.extend(page_parser.links)
            emails.extend(extract_emails(page_html + "\n" + page_parser.text + "\n" + page_text))
            phones.extend(extract_phones(page_html + "\n" + page_parser.text + "\n" + page_text))
            socials = merge_socials(socials, extract_social_links(page_parser.links))
            addresses = merge_arrays(
                addresses,
                extract_jsonld_addresses(page_jsonld),
                extract_address_candidates(page_text),
                extract_location_mentions(page_text),
            )
            if not summary:
                summary = clean_text(page_parser.meta_description)
            description = pick_longer(description, page_parser.text[:1200], minimum_gain=80)
            if not application_link:
                for link in page_parser.links:
                    href = normalize_website(link.get("href"))
                    haystack = f"{href or ''} {link.get('text') or ''}".lower()
                    if href and any(keyword in haystack for keyword in ("apply", "application", "grant", "funding", "partner")):
                        application_link = href
                        break
            time.sleep(0.3)
        known_work_links = extract_known_work_links(final_url, all_links)
        return {
            "final_url": final_url,
            "summary": clean_text(summary),
            "description": clean_text(description),
            "emails": dedupe(emails),
            "phones": dedupe(phones),
            "socials": socials,
            "known_work_links": known_work_links,
            "addresses": addresses,
            "application_link": application_link,
            "text": compact_spaces(" ".join(combined_text_parts)),
            "visited_pages": visited_pages,
            "notes": notes,
        }
    except Exception as exc:
        return {
            "final_url": row.website,
            "summary": None,
            "description": None,
            "emails": [],
            "phones": [],
            "socials": {},
            "known_work_links": [],
            "addresses": [],
            "application_link": None,
            "text": "",
            "visited_pages": [],
            "notes": [f"Website fetch failed: {exc}"],
        }


def build_type_specific_data(type_slug, row, enrichment, merged_description):
    workbook_geo = extract_region_tags(row.regional_presence)
    combined_text = " ".join([row.about, enrichment.get("text") or "", merged_description or ""])
    if type_slug == "cso":
        return {
            "areas_of_work": merge_arrays(row.thematic_areas),
            "beneficiary_groups": merge_arrays(infer_target_groups(combined_text)),
            "geography_served": merge_arrays(workbook_geo),
            "registration_status": infer_registration_status(combined_text),
            "registration_number": None,
            "programs": clean_text(first_sentence(merged_description or row.about, limit=500)),
            "volunteer_or_partner_needs": None,
        }
    if type_slug == "csr_philanthropy":
        return {
            "focus_areas": merge_arrays(row.thematic_areas),
            "geography_served": merge_arrays(workbook_geo),
            "support_instruments": infer_support_instruments(combined_text, row.org_type),
            "typical_support_size": None,
            "beneficiary_or_partner_focus": merge_arrays(infer_target_groups(combined_text)),
            "application_or_nomination_process": None,
            "application_link": enrichment.get("application_link"),
            "partnership_preferences": [],
            "reporting_or_compliance_notes": "FCRA eligibility relevant" if FCRA_RE.search(combined_text) else None,
        }
    socials = enrichment.get("socials") or {}
    return {
        "storytelling_modes": infer_storytelling_modes(combined_text, socials, enrichment.get("known_work_links") or []),
        "youtube_url": socials.get("youtube"),
        "instagram_url": socials.get("instagram"),
        "linkedin_url": socials.get("linkedin"),
        "facebook_url": socials.get("facebook"),
        "portfolio_website": enrichment.get("final_url") or row.website,
        "other_social_links": "\n".join(
            f"{platform.title()} | {url}"
            for platform, url in socials.items()
            if platform not in {"youtube", "instagram", "linkedin", "facebook"}
        ) or None,
        "geography_served": merge_arrays(workbook_geo),
        "languages": [],
        "reach": None,
        "target_audience": merge_arrays(infer_target_groups(combined_text)),
        "known_work_links": "\n".join(enrichment.get("known_work_links") or []) or None,
    }


def merge_type_specific(existing, incoming):
    merged = dict(existing or {})
    for key, value in (incoming or {}).items():
        if isinstance(value, list):
            merged[key] = merge_arrays(merged.get(key) or [], value)
        elif isinstance(value, str):
            merged[key] = pick_longer(merged.get(key), value, minimum_gain=12)
        else:
            merged[key] = value if value not in (None, "", []) else merged.get(key)
    return merged


def build_admin_notes(row, existing, enrichment, type_slug, main_address):
    notes = [
        f"Source workbook category: {row.org_type or 'Unknown'}",
        f"Regional presence: {row.regional_presence or 'Unknown'}",
        f"Imported on: 2026-05-10",
        f"Target type: {type_slug}",
    ]
    if existing:
        notes.append(f"Merged into existing entity_uid: {existing.get('entity_uid')}")
    if main_address:
        notes.append(f"Main office used for map: {main_address}")
    if enrichment.get("visited_pages"):
        notes.append("Reviewed pages: " + "; ".join(enrichment["visited_pages"][:5]))
    for note in enrichment.get("notes") or []:
        notes.append(note)
    return "\n".join(notes)


def build_record(row, existing, geocode_cache):
    desired_type = classify_row(row)
    manual_override = MANUAL_OVERRIDES.get(slugify(row.organization_name), {})
    if manual_override.get("entity_type_slug"):
        desired_type = manual_override["entity_type_slug"]
    enrichment = enrich_from_website(row)
    if row.org_type.lower() == "international orgnaizations":
        combined = " ".join([row.about, row.organization_name]).lower()
        desired_type = "csr_philanthropy" if re.search(r"\b(fund|foundation|grant|philanthrop|donor)\b", combined, re.I) else "cso"
    if manual_override.get("entity_type_slug"):
        desired_type = manual_override["entity_type_slug"]
    target_type = existing.get("entity_type_slug") if existing else desired_type

    website = enrichment.get("final_url") or row.website or (existing.get("website_url") if existing else None)
    socials = merge_socials(existing.get("social_media") if existing else {}, enrichment.get("socials"))
    summary = pick_longer(existing.get("summary") if existing else None, first_sentence(row.about), minimum_gain=0)
    summary = pick_longer(summary, enrichment.get("summary"), minimum_gain=8)
    description = pick_longer(existing.get("description") if existing else None, row.about, minimum_gain=20)
    description = pick_longer(description, enrichment.get("description"), minimum_gain=80)

    main_address = manual_override.get("primary_address") or pick_main_address(enrichment.get("addresses") or []) or (existing.get("primary_address") if existing else None)
    office_locations = merge_arrays(
        (existing.get("office_locations") if existing else []) or [],
        [item for item in (enrichment.get("addresses") or []) if item != main_address][0:6],
    )
    geography = merge_arrays(
        extract_region_tags(row.regional_presence),
        [existing.get("state")] if existing and existing.get("state") else [],
    )
    state = manual_override.get("state") or extract_state(main_address or "", geography) or (existing.get("state") if existing else None)
    district = manual_override.get("district") or extract_district(main_address or "", state) or (existing.get("district") if existing else None)
    location_label = manual_override.get("location_label") or ", ".join(part for part in [district, state] if clean_text(part)) or (
        existing.get("location_label") if existing else None
    ) or ", ".join(extract_region_tags(row.regional_presence))
    latitude = manual_override.get("latitude")
    longitude = manual_override.get("longitude")
    if latitude is None or longitude is None:
        latitude = existing.get("latitude") if existing else None
        longitude = existing.get("longitude") if existing else None
    if not has_usable_coordinate(latitude, longitude):
        latitude, longitude = geocode_entity(
            row.organization_name,
            main_address,
            location_label,
            state,
            geography,
            geocode_cache,
        )

    type_specific = build_type_specific_data(target_type, row, enrichment, description)
    if existing:
        type_specific = merge_type_specific(existing.get("type_specific_data") or {}, type_specific)

    tags = merge_arrays(
        (existing.get("tags") if existing else []) or [],
        row.thematic_areas,
        extract_region_tags(row.regional_presence),
        [row.org_type] if row.org_type else [],
        [SOURCE_LABEL],
    )
    keywords = merge_arrays(
        (existing.get("keywords") if existing else []) or [],
        row.thematic_areas,
        extract_region_tags(row.regional_presence),
        [row.org_type, row.organization_name],
    )
    admin_notes = build_admin_notes(row, existing, enrichment, target_type, main_address)
    entity_uid = existing.get("entity_uid") if existing else build_entity_uid(target_type, row.organization_name, website)
    record = {
        "table": TARGET_TABLES[target_type],
        "entity_type_slug": target_type,
        "entity_uid": entity_uid,
        "entity_name": row.organization_name,
        "summary": summary,
        "description": description,
        "location_label": location_label,
        "primary_address": main_address,
        "district": district,
        "state": state,
        "country": manual_override.get("country") or (existing.get("country") if existing and clean_text(existing.get("country")) else "India"),
        "contact_email": (enrichment.get("emails") or [None])[0] or (existing.get("contact_email") if existing else None),
        "contact_phone": (enrichment.get("phones") or [None])[0] or (existing.get("contact_phone") if existing else None),
        "website_url": website,
        "social_media": socials,
        "office_locations": office_locations,
        "tags": tags,
        "keywords": keywords,
        "latitude": latitude,
        "longitude": longitude,
        "source_label": SOURCE_LABEL,
        "source_url": website,
        "type_specific_data": type_specific,
        "created_by_name": (existing.get("created_by_name") if existing else None) or CREATED_BY_NAME,
        "created_by_email": (existing.get("created_by_email") if existing else None) or CREATED_BY_EMAIL,
        "admin_notes": admin_notes,
        "approval_status": "approved",
        "approved_at": "now()",
        "approved_by": "admin",
        "is_deleted": False,
        "updated_at": "now()",
        "search_text": None,
        "merge_action": "merged" if existing else "inserted",
        "website_pages_reviewed": enrichment.get("visited_pages") or [],
    }
    record["search_text"] = build_search_text(record)
    return record


def build_values_sql(record):
    return "(" + ", ".join([
        sql_text(record["entity_uid"]),
        sql_text(record["entity_name"]),
        sql_text(record["summary"]),
        sql_text(record["description"]),
        sql_text(record["location_label"]),
        sql_text(record["primary_address"]),
        sql_text(record["district"]),
        sql_text(record["state"]),
        sql_text(record["country"]),
        sql_text(record["contact_email"]),
        sql_text(record["contact_phone"]),
        sql_text(record["website_url"]),
        sql_json(record["social_media"]),
        sql_json(record["office_locations"]),
        sql_text_array(record["tags"]),
        sql_text_array(record["keywords"]),
        "null" if record["latitude"] in (None, "") else str(record["latitude"]),
        "null" if record["longitude"] in (None, "") else str(record["longitude"]),
        sql_text(record["source_label"]),
        sql_text(record["source_url"]),
        sql_json(record["type_specific_data"]),
        sql_text(record["created_by_name"]),
        sql_text(record["created_by_email"]),
        sql_text(record["admin_notes"]),
        "'approved'",
        "now()",
        "'admin'",
        "false",
        sql_text(record["search_text"]),
        "now()",
    ]) + ")"


def build_insert_block(type_slug, table, records):
    values_sql = ",\n".join(build_values_sql(record) for record in records)
    return f"""insert into public.{table} (
  entity_uid,
  entity_name,
  summary,
  description,
  location_label,
  primary_address,
  district,
  state,
  country,
  contact_email,
  contact_phone,
  website_url,
  social_media,
  office_locations,
  tags,
  keywords,
  latitude,
  longitude,
  source_label,
  source_url,
  type_specific_data,
  created_by_name,
  created_by_email,
  admin_notes,
  approval_status,
  approved_at,
  approved_by,
  is_deleted,
  search_text,
  updated_at
) values
{values_sql}
on conflict (entity_uid) do update set
  entity_name = excluded.entity_name,
  summary = excluded.summary,
  description = excluded.description,
  location_label = excluded.location_label,
  primary_address = excluded.primary_address,
  district = excluded.district,
  state = excluded.state,
  country = excluded.country,
  contact_email = excluded.contact_email,
  contact_phone = excluded.contact_phone,
  website_url = excluded.website_url,
  social_media = excluded.social_media,
  office_locations = excluded.office_locations,
  tags = excluded.tags,
  keywords = excluded.keywords,
  latitude = excluded.latitude,
  longitude = excluded.longitude,
  source_label = excluded.source_label,
  source_url = excluded.source_url,
  type_specific_data = excluded.type_specific_data,
  created_by_name = excluded.created_by_name,
  created_by_email = excluded.created_by_email,
  admin_notes = excluded.admin_notes,
  approval_status = 'approved',
  approved_at = now(),
  approved_by = 'admin',
  is_deleted = false,
  search_text = excluded.search_text,
  updated_at = now();
"""


def main():
    base_url, anon_key = read_config()
    workbook_rows = load_workbook_rows()
    existing_entities = fetch_existing_entities(base_url, anon_key)
    geocode_cache = {}
    records_by_uid = {}
    match_pool = list(existing_entities)
    report = {
        "source_label": SOURCE_LABEL,
        "total_rows": len(workbook_rows),
        "by_type": {type_slug: 0 for type_slug in TARGET_TABLES},
        "merged_rows": 0,
        "inserted_rows": 0,
        "missing_coordinates": [],
    }

    for index, row in enumerate(workbook_rows, start=1):
        desired_type = classify_row(row)
        existing = choose_existing_match(row, desired_type, match_pool)
        record = build_record(row, existing, geocode_cache)
        report[f"{record['merge_action']}_rows"] += 1
        records_by_uid[record["entity_uid"]] = record
        match_pool = [
            item for item in match_pool
            if item.get("entity_uid") != record["entity_uid"]
        ]
        match_pool.append(record)
        if not has_usable_coordinate(record["latitude"], record["longitude"]):
            report["missing_coordinates"].append({
                "entity_name": record["entity_name"],
                "entity_type_slug": record["entity_type_slug"],
                "website_url": record["website_url"],
                "location_label": record["location_label"],
                "primary_address": record["primary_address"],
            })
        print(f"[{index}/{len(workbook_rows)}] {record['entity_name']} -> {record['entity_type_slug']} ({record['merge_action']})")
        time.sleep(0.25)

    grouped_records = {type_slug: [] for type_slug in TARGET_TABLES}
    deduped_missing = []
    missing_seen = set()
    for record in records_by_uid.values():
        grouped_records[record["entity_type_slug"]].append(record)
        report["by_type"][record["entity_type_slug"]] += 1
        if not has_usable_coordinate(record["latitude"], record["longitude"]):
            key = record["entity_uid"]
            if key not in missing_seen:
                missing_seen.add(key)
                deduped_missing.append({
                    "entity_name": record["entity_name"],
                    "entity_type_slug": record["entity_type_slug"],
                    "website_url": record["website_url"],
                    "location_label": record["location_label"],
                    "primary_address": record["primary_address"],
                })
    report["missing_coordinates"] = deduped_missing

    sql_blocks = [
        build_insert_block(type_slug, TARGET_TABLES[type_slug], records)
        for type_slug, records in grouped_records.items()
        if records
    ]
    OUTPUT_SQL_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL_PATH.write_text("\n\n".join(sql_blocks) + "\n", encoding="utf-8")
    OUTPUT_REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote SQL migration to {OUTPUT_SQL_PATH}")
    print(f"Wrote import report to {OUTPUT_REPORT_PATH}")
    print(json.dumps(report, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
