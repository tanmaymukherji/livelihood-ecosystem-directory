import hashlib
import json
import re
import ssl
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\github\livelihood-ecosystem-directory")
WORKBOOK_PATH = Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\CNH Infrastructure_May 2026.xlsx")
OUTPUT_SQL_PATH = PROJECT_ROOT / "supabase" / "migrations" / "20260510170000_import_cnh_story_tellers.sql"
OUTPUT_REPORT_PATH = PROJECT_ROOT / "scripts" / "cnh_story_teller_import_report.json"
CONFIG_PATH = PROJECT_ROOT / "config.js"
PUBLIC_VIEW = "ecosystem_directory_entities"

SOURCE_LABEL = "CNH Infrastructure May 2026 workbook"
CREATED_BY_NAME = "CNH story teller import"
CREATED_BY_EMAIL = "tanmay@greenruraleconomy.in"
INDIA_CENTER = (22.3511148, 78.6677428)
USER_AGENT = "Livelihood Ecosystem Directory CNH Story Teller Import/1.0"
SSL_CONTEXT = ssl._create_unverified_context()


@dataclass
class CreatorRow:
    name: str
    link: str
    language: str
    state: str


@dataclass
class PrintRow:
    name: str
    publication: str


PUBLICATION_DEFAULTS = {
    "Agence France-Presse (AFP)": {
        "website_url": "https://www.afp.com/en",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "geography_served": ["India", "Global"],
    },
    "PTI": {
        "website_url": "https://www.ptinews.com/",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "geography_served": ["India"],
    },
    "Bloomberg": {
        "website_url": "https://www.bloomberg.com/",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "geography_served": ["India", "Global"],
    },
    "IANS": {
        "website_url": "https://ians.in/",
        "location_label": "Chandigarh, India",
        "primary_address": "Chandigarh, India",
        "geography_served": ["India"],
    },
    "Reuters": {
        "website_url": "https://www.reuters.com/",
        "location_label": "Chennai, Tamil Nadu",
        "primary_address": "Chennai, Tamil Nadu, India",
        "geography_served": ["India", "Global"],
    },
    "Times of India": {
        "website_url": "https://timesofindia.indiatimes.com/",
        "location_label": "India",
        "primary_address": "India",
        "geography_served": ["India"],
    },
    "Hindustan Times": {
        "website_url": "https://www.hindustantimes.com/",
        "location_label": "India",
        "primary_address": "India",
        "geography_served": ["India"],
    },
    "The Hindu": {
        "website_url": "https://www.thehindu.com/",
        "location_label": "India",
        "primary_address": "India",
        "geography_served": ["India"],
    },
    "The Pioneer": {
        "website_url": "https://www.dailypioneer.com/",
        "location_label": "India",
        "primary_address": "India",
        "geography_served": ["India"],
    },
    "The Indian Express": {
        "website_url": "https://indianexpress.com/",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "geography_served": ["India"],
    },
    "News18": {
        "website_url": "https://www.news18.com/",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "geography_served": ["India"],
    },
    "Mid-Day": {
        "website_url": "https://www.mid-day.com/",
        "location_label": "Mumbai, Maharashtra",
        "primary_address": "Mumbai, Maharashtra, India",
        "geography_served": ["Mumbai", "India"],
    },
    "The Wire": {
        "website_url": "https://thewire.in/",
        "location_label": "India",
        "primary_address": "India",
        "geography_served": ["India"],
    },
}


PRINT_CONTACT_METADATA = {
    "Uzmi Athar": {
        "summary": "South Asia correspondent with Agence France-Presse (AFP), reporting on climate, biodiversity, pollution, energy, forests, and social-development issues.",
        "description": "New Delhi-based journalist working as a South Asia correspondent with AFP. Her reporting spans climate change, biodiversity, energy, forests, wildlife trafficking, and social-development issues across India and the region.",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "known_work_links": [
            "https://muckrack.com/uzmi-athar",
            "https://earthjournalism.net/fellows/uzmi-athar",
            "https://in.linkedin.com/in/uzmi-athar",
        ],
        "other_social_links": ["LinkedIn | https://in.linkedin.com/in/uzmi-athar"],
        "target_audience": ["Policy", "Climate", "Environment"],
    },
    "Gaurav Saini": {
        "summary": "Senior PTI reporter covering climate change, environment policy, agriculture, wildlife, water, and urban issues.",
        "description": "Delhi-based multimedia reporter with PTI covering environment policy and climate change since 2019, with prior reporting across governance, politics, and urban planning.",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "known_work_links": [
            "https://climatereporting.stanleycenter.org/pti-journalists/",
            "https://climatereporting.stanleycenter.org/cop28-reflections/",
            "https://climatereporting.stanleycenter.org/cop28-fellowship/",
        ],
        "target_audience": ["Policy", "Climate", "Environment", "Agriculture"],
    },
    "Pratik Parija": {
        "summary": "Bloomberg reporter covering climate, agriculture, commodities, and energy from India.",
        "description": "Reporter for Bloomberg with bylines on agriculture, commodities, monsoon, fertilizer, and energy-market issues affecting India.",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "known_work_links": [
            "https://www.bloomberg.com/authors/ACT1hUlVowE/pratik-parija",
            "https://www.bloomberg.com/news/articles/2026-03-09/india-farmers-stage-fresh-protest-in-delhi-against-us-trade-deal",
        ],
        "target_audience": ["Agriculture", "Climate", "Markets"],
    },
    "Vishal Gulati": {
        "summary": "IANS journalist covering biodiversity, climate change, and environment-development issues.",
        "description": "Chandigarh-based journalist with IANS covering biodiversity, climate change, and the links between environment and development.",
        "location_label": "Chandigarh, India",
        "primary_address": "Chandigarh, India",
        "known_work_links": [
            "https://www.cleanenergywire.org/members/gulati",
        ],
        "other_social_links": ["X | https://x.com/VishalG1010"],
        "target_audience": ["Climate", "Environment", "Biodiversity"],
    },
    "Sudarshan Varadhan": {
        "summary": "Reuters journalist covering energy, environment, climate, and rights issues from Chennai.",
        "description": "Reuters journalist based in Chennai with reporting across energy, environment, rights, and regional issues in India and Asia.",
        "location_label": "Chennai, Tamil Nadu",
        "primary_address": "Chennai, Tamil Nadu, India",
        "contact_email": "Sudarshan.varadhan@tr.com",
        "known_work_links": [
            "https://muckrack.com/sudarshan-varadhan",
            "https://www.reuters.com/world/africa/south-africa-boosts-coal-exports-israel-after-colombia-ban-2025-12-16/",
        ],
        "target_audience": ["Energy", "Climate", "Environment"],
    },
    "Kushagra Dixit": {
        "summary": "Times of India journalist writing on wildlife, climate change, agriculture, human rights, and scientific research.",
        "description": "Times of India journalist whose work spans wildlife conservation, climate change, agriculture, air pollution, urban waste, human rights, and scientific research.",
        "location_label": "India",
        "primary_address": "India",
        "known_work_links": [
            "https://timesofindia.indiatimes.com/toireportergadgetposts.cms?author=Kushagra-Dixit&authorid=479256218&from=mdr",
        ],
        "target_audience": ["Climate", "Wildlife", "Agriculture", "Science"],
    },
    "Richa Pinto": {
        "summary": "Times of India special correspondent covering urban governance and climate issues in Mumbai.",
        "description": "Special correspondent with The Times of India. She covers urban governance and climate change issues, with extensive reporting on civic issues affecting Mumbai.",
        "location_label": "Mumbai, Maharashtra",
        "primary_address": "Mumbai, Maharashtra, India",
        "known_work_links": [
            "https://timesofindia.indiatimes.com/toireporter/author-Richa-Pinto-479196857.cms",
            "https://muckrack.com/richa-pinto",
        ],
        "other_social_links": ["X | https://x.com/richapintoi"],
        "target_audience": ["Cities", "Climate", "Urban Governance"],
    },
    "Sanjay Dutta": {
        "summary": "Times of India journalist and editor with broad reporting on development and public-affairs issues.",
        "description": "Journalist with The Times of India working on public-interest reporting and editorial coverage in India.",
        "location_label": "India",
        "primary_address": "India",
        "known_work_links": [
            "https://timesofindia.indiatimes.com/",
        ],
        "target_audience": ["Public Affairs", "Development"],
    },
    "Abhishek Jha": {
        "summary": "Hindustan Times data journalist writing on climate and weather trends.",
        "description": "Data journalist at Hindustan Times whose recent reporting includes monsoon and climate analysis in India.",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "known_work_links": [
            "https://www.hindustantimes.com/environment/weather-bee-is-the-monsoon-becoming-rainier-101723728641284.html",
            "https://journalists.feedspot.com/hindustan_times_journalists/",
        ],
        "other_social_links": ["X | https://x.com/naalmot"],
        "target_audience": ["Climate", "Data", "Weather"],
    },
    "Abhishek Behl": {
        "summary": "Hindustan Times journalist covering urban development, infrastructure, real estate, and civic issues in Gurugram.",
        "description": "Hindustan Times reporter focused on urban development, real estate, transport, and infrastructure in Gurugram.",
        "location_label": "Gurugram, Haryana",
        "primary_address": "Gurugram, Haryana, India",
        "contact_email": "abhishek.behl@hindustantimes.com",
        "known_work_links": [
            "https://stg-www.hindustantimes.com/cities/gurugram-news/gurugram-water-supply-disrupted-for-third-day-after-gmda-pipeline-damage-101768156201248-amp.html",
            "https://x.com/abhishek_behl",
        ],
        "other_social_links": ["X | https://x.com/abhishek_behl"],
        "target_audience": ["Urban Development", "Infrastructure", "Real Estate"],
    },
    "Jayashree Nandi": {
        "summary": "Hindustan Times environment and climate correspondent covering climate science, negotiations, weather, and environmental policy.",
        "description": "Environment and climate correspondent at Hindustan Times. She covers climate science, UN climate negotiations, weather, environmental policy, litigation, and people’s stories.",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "known_work_links": [
            "https://pulitzercenter.org/people/jayashree-nandi",
            "https://rainforestjournalismfund.org/people/jayashree-nandi",
            "https://journalistdb.com/journalists/jayashree-nandi/",
        ],
        "target_audience": ["Climate", "Environment", "Policy"],
    },
    "Jasjeev Gandhiok": {
        "summary": "Hindustan Times journalist reporting on climate change, water, and environmental issues from Delhi.",
        "description": "Journalist at Hindustan Times with coverage focused on climate change, water, and environmental and civic issues in and around Delhi.",
        "location_label": "Delhi, India",
        "primary_address": "Delhi, India",
        "known_work_links": [
            "https://intelligentrelations.com/journalist/jasjeev-gandhiok/",
        ],
        "target_audience": ["Climate", "Environment", "Cities"],
    },
    "K C Deepika": {
        "summary": "City Editor at The Hindu, Bengaluru, with reporting across civic, education, power, tourism, and environment issues.",
        "description": "City Editor with The Hindu in Bengaluru. She has covered a wide range of issues from civic and education to power, tourism, and environment.",
        "location_label": "Bengaluru, Karnataka",
        "primary_address": "Bengaluru, Karnataka, India",
        "known_work_links": [
            "https://buzzsumo.com/journalist/k-c-deepika-30797445/",
            "https://azimpremjiuniversity.edu.in/news/2024/getting-to-the-root-of-bengalurus-tree-falls",
        ],
        "other_social_links": ["X | https://x.com/deepikacariappa"],
        "target_audience": ["Cities", "Environment", "Public Affairs"],
    },
    "Nikhil M Babu": {
        "summary": "The Hindu journalist covering environment, governance, and politics.",
        "description": "Journalist with The Hindu writing on environment, governance, and politics, with reporting that has included Delhi climate and governance issues.",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "known_work_links": [
            "https://buzzsumo.com/journalist/nikhil-m-babu-131274189/",
        ],
        "other_social_links": ["X | https://x.com/ikasnik"],
        "target_audience": ["Environment", "Governance", "Politics"],
    },
    "B. Aravind Kumar": {
        "summary": "Deputy Bureau Chief at The Hindu, Chennai, covering environment, current affairs, and city reporting.",
        "description": "Deputy Bureau Chief with The Hindu based in Chennai, with coverage across environment, city reporting, and current affairs.",
        "location_label": "Chennai, Tamil Nadu",
        "primary_address": "Chennai, Tamil Nadu, India",
        "known_work_links": [
            "https://www.goskribe.com/profile/b-aravind-kumar",
        ],
        "target_audience": ["Environment", "Cities", "Current Affairs"],
    },
    "Kota Sriraj": {
        "summary": "Journalist and commentator writing on environment and biodiversity issues.",
        "description": "Environmental journalist and commentator with public writing on biodiversity and ecology-related issues.",
        "location_label": "India",
        "primary_address": "India",
        "known_work_links": [
            "https://www.newsofbahrain.com/views/68153.html",
        ],
        "target_audience": ["Environment", "Biodiversity"],
    },
    "Amitabh Sinha": {
        "summary": "Deputy Editor at The Indian Express writing on environment, climate change, water, science, and technology.",
        "description": "Deputy Editor at The Indian Express. He writes on environment, climate change, water, science and technology, and has previously worked with PTI, Reuters, and BBC.",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "known_work_links": [
            "https://indianexpress.com/profile/columnist/amitabh-sinha/",
            "https://wmo.int/profile/amitabh-sinha",
        ],
        "target_audience": ["Climate", "Science", "Policy"],
    },
    "Srishti Choudhary": {
        "summary": "CNN-News18 science and environment journalist covering climate change, politics, science, and weather.",
        "description": "Science and environment journalist with CNN-News18 based in New Delhi, covering climate change, science, politics, and weather.",
        "location_label": "New Delhi, India",
        "primary_address": "New Delhi, India",
        "known_work_links": [
            "https://muckrack.com/srishti-choudhary",
            "https://www.helpareporter.com/journalist/srishti-choudhary",
        ],
        "other_social_links": ["Instagram | https://www.instagram.com/khair.khabar"],
        "target_audience": ["Climate", "Science", "Environment"],
    },
    "Ranjeet Jadhav": {
        "summary": "Mid-Day environment journalist and assistant editor focused on wildlife, conservation, and urban ecology in Mumbai.",
        "description": "Assistant editor and journalist with Mid-Day in Mumbai, known for environment and wildlife coverage and award-winning reporting on Aarey and conservation issues.",
        "location_label": "Mumbai, Maharashtra",
        "primary_address": "Mumbai, Maharashtra, India",
        "known_work_links": [
            "https://muckrack.com/ranjeet-jadhav",
            "https://www.mid-day.com/news/india-news/article/mid-day-journalist-ranjeet-jadhav-wins-redink-star-mumbai-reporter-award-21245164",
        ],
        "target_audience": ["Environment", "Wildlife", "Cities"],
    },
    "Aathira Perinchery": {
        "summary": "The Wire reporter focused on ecology, conservation, and wildlife research in India.",
        "description": "A trained wildlife biologist turned journalist whose reporting for The Wire focuses on ecology, conservation, and wildlife research in India.",
        "location_label": "India",
        "primary_address": "India",
        "known_work_links": [
            "https://reportfortheworld.org/members/aathira-perinchery/",
            "https://earthjournalism.net/stories/the-hidden-ravages-of-the-poultry-boom-in-tamil-nadu-india",
        ],
        "target_audience": ["Ecology", "Conservation", "Wildlife"],
    },
    "Geetha Srimathi": {
        "summary": "The Hindu reporter in Chennai covering environment and civic issues.",
        "description": "Reporter with The Hindu based in Chennai, with public bylines on environment, civic issues, and urban change in Tamil Nadu.",
        "location_label": "Chennai, Tamil Nadu",
        "primary_address": "Chennai, Tamil Nadu, India",
        "known_work_links": [
            "https://muckrack.com/geetha-srimathi",
            "https://www.goskribe.com/profile/geetha-srimathi",
        ],
        "target_audience": ["Environment", "Cities", "Civic Issues"],
    },
}


def compact_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def clean_text(value):
    text = compact_spaces(value)
    if not text or text.lower() in {"na", "n/a", "none", "null", "-"}:
        return None
    return text


def slugify(value: str) -> str:
    value = re.sub(r"[^\w\s-]", "", str(value or "").strip().lower())
    value = re.sub(r"[-\s]+", "-", value)
    return value.strip("-")


def dedupe(values):
    seen = set()
    output = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        output.append(text)
    return output


def sql_text(value):
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_text_array(values):
    cleaned = [str(item).replace("'", "''") for item in values if clean_text(item)]
    if not cleaned:
        return "'{}'::text[]"
    return "ARRAY[" + ", ".join(f"'{item}'" for item in cleaned) + "]::text[]"


def sql_json(value):
    return sql_text(json.dumps(value, ensure_ascii=False)) + "::jsonb"


def read_config():
    text = CONFIG_PATH.read_text(encoding="utf-8")
    url_match = re.search(r"SUPABASE_URL:\s*'([^']+)'", text)
    key_match = re.search(r"SUPABASE_ANON_KEY:\s*'([^']+)'", text)
    if not url_match or not key_match:
        raise RuntimeError("Could not read Supabase config from config.js")
    return url_match.group(1), key_match.group(1)


def fetch_existing_story_tellers(base_url, anon_key):
    query = ",".join([
        "entity_uid",
        "entity_name",
        "summary",
        "description",
        "location_label",
        "primary_address",
        "district",
        "state",
        "country",
        "contact_email",
        "contact_phone",
        "website_url",
        "social_media",
        "office_locations",
        "tags",
        "keywords",
        "latitude",
        "longitude",
        "source_label",
        "source_url",
        "type_specific_data",
        "created_by_name",
        "created_by_email",
        "admin_notes",
        "search_text",
    ])
    url = (
        f"{base_url}/rest/v1/{PUBLIC_VIEW}"
        f"?select={urllib.parse.quote(query)}"
        f"&entity_type_slug=eq.story_teller&limit=1000"
    )
    headers = {
        "apikey": anon_key,
        "Authorization": f"Bearer {anon_key}",
        "Accept": "application/json",
    }
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
        return json.loads(response.read().decode("utf-8"))


def load_workbook_rows():
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    creators = []
    contacts = []
    creator_ws = workbook["Creator List"]
    for row in creator_ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        name = clean_text(row[1])
        if not name:
            continue
        creators.append(CreatorRow(
            name=name,
            link=clean_text(row[2]) or "",
            language=clean_text(row[3]) or "",
            state=clean_text(row[4]) or "",
        ))
    contact_ws = workbook["PR & Publications"]
    for row in contact_ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        name = clean_text(row[1])
        publication = clean_text(row[2])
        if not name or not publication:
            continue
        contacts.append(PrintRow(name=name, publication=publication))
    return creators, contacts


def merge_arrays(*arrays):
    merged = []
    for array in arrays:
        if isinstance(array, list):
            merged.extend(array)
    return dedupe(merged)


def merge_socials(*items):
    output = {}
    for item in items:
        if isinstance(item, dict):
            for key, value in item.items():
                clean = clean_text(value)
                if clean and key not in output:
                    output[key] = clean
    return output


def pick_text(primary, secondary):
    return clean_text(primary) or clean_text(secondary)


def build_entity_uid(name: str):
    digest = hashlib.md5(name.encode("utf-8")).hexdigest()[:8]
    return f"story_teller-{slugify(name)}-{digest}"


def normalize_instagram_profile(link):
    text = clean_text(link)
    if not text or "instagram.com" not in text.lower():
        return None
    if "/reel/" in text or "/p/" in text or "/tv/" in text:
        return None
    return text


def infer_creator_modes(link):
    lowered = (link or "").lower()
    if "instagram.com" in lowered:
        return ["Social Media", "Video"]
    if "youtu" in lowered:
        return ["Video", "Audio", "Social Media"]
    return ["Social Media"]


def infer_creator_platform_fields(link):
    lowered = (link or "").lower()
    if "instagram.com" in lowered:
        return {
            "instagram_url": normalize_instagram_profile(link),
            "youtube_url": None,
            "other_social_links": None,
            "social_media": {"instagram": link},
        }
    if "youtu" in lowered:
        return {
            "instagram_url": None,
            "youtube_url": link,
            "other_social_links": None,
            "social_media": {"youtube": link},
        }
    return {
        "instagram_url": None,
        "youtube_url": None,
        "other_social_links": f"Link | {link}" if link else None,
        "social_media": {"link": link} if link else {},
    }


def geocode_query(query, cache):
    query = clean_text(query)
    if not query:
        return None
    if query in cache:
        return cache[query]
    url = f"https://nominatim.openstreetmap.org/search?format=jsonv2&limit=1&q={urllib.parse.quote(query)}"
    request = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception:
        cache[query] = None
        return None
    match = data[0] if isinstance(data, list) and data else None
    if not match:
        cache[query] = None
        return None
    try:
        point = (float(match["lat"]), float(match["lon"]))
    except (TypeError, KeyError, ValueError):
        cache[query] = None
        return None
    cache[query] = point
    return point


def geocode_location(location_label, primary_address, state, cache):
    queries = dedupe([
        primary_address,
        location_label,
        ", ".join(part for part in [state, "India"] if clean_text(part)),
        "India",
    ])
    for query in queries:
        point = geocode_query(query, cache)
        if point:
            if query != "India":
                time.sleep(0.8)
            return point
    return INDIA_CENTER


def build_search_text(record):
    values = [
        record.get("entity_name"),
        record.get("summary"),
        record.get("description"),
        record.get("location_label"),
        record.get("primary_address"),
        record.get("district"),
        record.get("state"),
        record.get("country"),
        " ".join(record.get("tags") or []),
        " ".join(record.get("keywords") or []),
    ]
    for value in (record.get("type_specific_data") or {}).values():
        if isinstance(value, list):
            values.append(" ".join(str(item) for item in value if clean_text(item)))
        else:
            values.append(clean_text(value) or "")
    return compact_spaces(" ".join(item for item in values if clean_text(item)))


def build_creator_record(row: CreatorRow, existing, geocode_cache):
    platform_fields = infer_creator_platform_fields(row.link)
    location_label = f"{row.state}, India" if row.state else "India"
    primary_address = location_label
    lat, lng = geocode_location(location_label, primary_address, row.state, geocode_cache)
    summary = f"Story teller creating primarily in {row.language} from {row.state}." if row.language and row.state else f"Story teller listed in the CNH creator workbook."
    description = f"Listed in the CNH Infrastructure May 2026 creator list with a public {('Instagram' if 'instagram.com' in row.link.lower() else 'YouTube' if 'youtu' in row.link.lower() else 'social media')} link."
    type_specific_data = {
        "storytelling_modes": infer_creator_modes(row.link),
        "youtube_url": platform_fields["youtube_url"],
        "instagram_url": platform_fields["instagram_url"],
        "linkedin_url": None,
        "facebook_url": None,
        "portfolio_website": None,
        "other_social_links": platform_fields["other_social_links"],
        "geography_served": [row.state] if row.state else ["India"],
        "languages": [row.language] if row.language else [],
        "reach": None,
        "target_audience": [],
        "known_work_links": row.link or None,
    }
    existing_type_specific = (existing or {}).get("type_specific_data") or {}
    merged_type_specific = dict(existing_type_specific)
    for key, value in type_specific_data.items():
        if isinstance(value, list):
            merged_type_specific[key] = merge_arrays(existing_type_specific.get(key) or [], value)
        else:
            merged_type_specific[key] = pick_text(existing_type_specific.get(key), value) or pick_text(value, existing_type_specific.get(key))
    entity_uid = (existing or {}).get("entity_uid") or build_entity_uid(row.name)
    record = {
        "entity_uid": entity_uid,
        "entity_name": row.name,
        "summary": pick_text((existing or {}).get("summary"), summary),
        "description": pick_text((existing or {}).get("description"), description),
        "location_label": pick_text((existing or {}).get("location_label"), location_label),
        "primary_address": pick_text((existing or {}).get("primary_address"), primary_address),
        "district": pick_text((existing or {}).get("district"), row.state),
        "state": pick_text((existing or {}).get("state"), row.state),
        "country": pick_text((existing or {}).get("country"), "India"),
        "contact_email": (existing or {}).get("contact_email"),
        "contact_phone": (existing or {}).get("contact_phone"),
        "website_url": pick_text((existing or {}).get("website_url"), platform_fields["instagram_url"] or platform_fields["youtube_url"]),
        "social_media": merge_socials((existing or {}).get("social_media"), platform_fields["social_media"]),
        "office_locations": merge_arrays((existing or {}).get("office_locations") or []),
        "tags": merge_arrays((existing or {}).get("tags") or [], ["Creator List", row.language, row.state, SOURCE_LABEL]),
        "keywords": merge_arrays((existing or {}).get("keywords") or [], [row.language, row.state, "creator", "story teller"]),
        "latitude": (existing or {}).get("latitude") or lat,
        "longitude": (existing or {}).get("longitude") or lng,
        "source_label": SOURCE_LABEL,
        "source_url": row.link or None,
        "type_specific_data": merged_type_specific,
        "created_by_name": (existing or {}).get("created_by_name") or CREATED_BY_NAME,
        "created_by_email": (existing or {}).get("created_by_email") or CREATED_BY_EMAIL,
        "admin_notes": f"Imported from CNH Infrastructure May 2026 workbook - Creator List",
        "approval_status": "approved",
        "approved_at": "now()",
        "approved_by": "admin",
        "is_deleted": False,
        "updated_at": "now()",
        "search_text": None,
        "merge_action": "merged" if existing else "inserted",
        "workbook_section": "Creator List",
    }
    record["search_text"] = build_search_text(record)
    return record


def build_print_record(row: PrintRow, existing, geocode_cache):
    publication = PUBLICATION_DEFAULTS.get(row.publication, {
        "website_url": None,
        "location_label": "India",
        "primary_address": "India",
        "geography_served": ["India"],
    })
    profile = PRINT_CONTACT_METADATA.get(row.name, {})
    location_label = pick_text(profile.get("location_label"), publication.get("location_label")) or "India"
    primary_address = pick_text(profile.get("primary_address"), publication.get("primary_address")) or "India"
    state = None
    if "," in location_label:
        parts = [clean_text(item) for item in location_label.split(",") if clean_text(item)]
        if len(parts) >= 2:
            state = parts[-1].replace("India", "").strip(" ,") or None
    lat, lng = geocode_location(location_label, primary_address, state, geocode_cache)
    publication_url = profile.get("publication_url") or publication.get("website_url")
    other_social_links = profile.get("other_social_links") or []
    social_media = {}
    instagram_url = None
    linkedin_url = None
    facebook_url = None
    youtube_url = None
    other_link_lines = []
    for item in other_social_links:
        clean = clean_text(item)
        if not clean:
            continue
        lower = clean.lower()
        if lower.startswith("instagram |"):
            instagram_url = clean.split("|", 1)[1].strip()
            social_media["instagram"] = instagram_url
        elif lower.startswith("linkedin |"):
            linkedin_url = clean.split("|", 1)[1].strip()
            social_media["linkedin"] = linkedin_url
        elif lower.startswith("facebook |"):
            facebook_url = clean.split("|", 1)[1].strip()
            social_media["facebook"] = facebook_url
        elif lower.startswith("youtube |"):
            youtube_url = clean.split("|", 1)[1].strip()
            social_media["youtube"] = youtube_url
        else:
            label, _, url = clean.partition("|")
            if url.strip():
                social_media[label.strip().lower()] = url.strip()
            other_link_lines.append(clean)
    known_work_links = dedupe(profile.get("known_work_links") or [])
    languages = ["English"]
    type_specific_data = {
        "storytelling_modes": ["Written"],
        "youtube_url": youtube_url,
        "instagram_url": instagram_url,
        "linkedin_url": linkedin_url,
        "facebook_url": facebook_url,
        "portfolio_website": publication_url,
        "other_social_links": "\n".join(other_link_lines) or None,
        "geography_served": publication.get("geography_served") or ["India"],
        "languages": languages,
        "reach": None,
        "target_audience": profile.get("target_audience") or [],
        "known_work_links": "\n".join(known_work_links) or publication_url,
    }
    existing_type_specific = (existing or {}).get("type_specific_data") or {}
    merged_type_specific = dict(existing_type_specific)
    for key, value in type_specific_data.items():
        if isinstance(value, list):
            merged_type_specific[key] = merge_arrays(existing_type_specific.get(key) or [], value)
        else:
            merged_type_specific[key] = pick_text(existing_type_specific.get(key), value) or pick_text(value, existing_type_specific.get(key))
    entity_uid = (existing or {}).get("entity_uid") or build_entity_uid(row.name)
    record = {
        "entity_uid": entity_uid,
        "entity_name": row.name,
        "summary": pick_text((existing or {}).get("summary"), profile.get("summary")) or f"Print media journalist at {row.publication}.",
        "description": pick_text((existing or {}).get("description"), profile.get("description")) or f"Journalist listed in the CNH PR & Publications sheet with publication {row.publication}.",
        "location_label": pick_text((existing or {}).get("location_label"), location_label),
        "primary_address": pick_text((existing or {}).get("primary_address"), primary_address),
        "district": (existing or {}).get("district"),
        "state": pick_text((existing or {}).get("state"), state),
        "country": pick_text((existing or {}).get("country"), "India"),
        "contact_email": pick_text((existing or {}).get("contact_email"), profile.get("contact_email")),
        "contact_phone": pick_text((existing or {}).get("contact_phone"), profile.get("contact_phone")),
        "website_url": pick_text((existing or {}).get("website_url"), publication_url),
        "social_media": merge_socials((existing or {}).get("social_media"), social_media),
        "office_locations": merge_arrays((existing or {}).get("office_locations") or []),
        "tags": merge_arrays((existing or {}).get("tags") or [], [row.publication, SOURCE_LABEL, "Print media"]),
        "keywords": merge_arrays((existing or {}).get("keywords") or [], [row.publication, "journalist", "print media"]),
        "latitude": (existing or {}).get("latitude") or lat,
        "longitude": (existing or {}).get("longitude") or lng,
        "source_label": SOURCE_LABEL,
        "source_url": publication_url,
        "type_specific_data": merged_type_specific,
        "created_by_name": (existing or {}).get("created_by_name") or CREATED_BY_NAME,
        "created_by_email": (existing or {}).get("created_by_email") or CREATED_BY_EMAIL,
        "admin_notes": f"Imported from CNH Infrastructure May 2026 workbook - PR & Publications",
        "approval_status": "approved",
        "approved_at": "now()",
        "approved_by": "admin",
        "is_deleted": False,
        "updated_at": "now()",
        "search_text": None,
        "merge_action": "merged" if existing else "inserted",
        "workbook_section": "PR & Publications",
    }
    record["search_text"] = build_search_text(record)
    return record


def build_values_sql(record):
    return "(" + ", ".join([
        sql_text(record["entity_uid"]),
        sql_text(record["entity_name"]),
        sql_text(record["summary"]),
        sql_text(record["description"]),
        sql_text(record["location_label"]),
        sql_text(record["primary_address"]),
        sql_text(record["district"]),
        sql_text(record["state"]),
        sql_text(record["country"]),
        sql_text(record["contact_email"]),
        sql_text(record["contact_phone"]),
        sql_text(record["website_url"]),
        sql_json(record["social_media"]),
        sql_json(record["office_locations"]),
        sql_text_array(record["tags"]),
        sql_text_array(record["keywords"]),
        str(record["latitude"]) if record["latitude"] is not None else "null",
        str(record["longitude"]) if record["longitude"] is not None else "null",
        sql_text(record["source_label"]),
        sql_text(record["source_url"]),
        sql_json(record["type_specific_data"]),
        sql_text(record["created_by_name"]),
        sql_text(record["created_by_email"]),
        sql_text(record["admin_notes"]),
        "'approved'",
        "now()",
        "'admin'",
        "false",
        sql_text(record["search_text"]),
        "now()",
    ]) + ")"


def build_insert_sql(records):
    return """insert into public.story_teller_entities (
  entity_uid,
  entity_name,
  summary,
  description,
  location_label,
  primary_address,
  district,
  state,
  country,
  contact_email,
  contact_phone,
  website_url,
  social_media,
  office_locations,
  tags,
  keywords,
  latitude,
  longitude,
  source_label,
  source_url,
  type_specific_data,
  created_by_name,
  created_by_email,
  admin_notes,
  approval_status,
  approved_at,
  approved_by,
  is_deleted,
  search_text,
  updated_at
) values
""" + ",\n".join(build_values_sql(record) for record in records) + """
on conflict (entity_uid) do update set
  entity_name = excluded.entity_name,
  summary = excluded.summary,
  description = excluded.description,
  location_label = excluded.location_label,
  primary_address = excluded.primary_address,
  district = excluded.district,
  state = excluded.state,
  country = excluded.country,
  contact_email = excluded.contact_email,
  contact_phone = excluded.contact_phone,
  website_url = excluded.website_url,
  social_media = excluded.social_media,
  office_locations = excluded.office_locations,
  tags = excluded.tags,
  keywords = excluded.keywords,
  latitude = excluded.latitude,
  longitude = excluded.longitude,
  source_label = excluded.source_label,
  source_url = excluded.source_url,
  type_specific_data = excluded.type_specific_data,
  created_by_name = excluded.created_by_name,
  created_by_email = excluded.created_by_email,
  admin_notes = excluded.admin_notes,
  approval_status = 'approved',
  approved_at = now(),
  approved_by = 'admin',
  is_deleted = false,
  search_text = excluded.search_text,
  updated_at = now();
"""


def main():
    base_url, anon_key = read_config()
    existing_rows = fetch_existing_story_tellers(base_url, anon_key)
    existing_by_name = {slugify(item.get("entity_name")): item for item in existing_rows}
    creators, contacts = load_workbook_rows()
    geocode_cache = {}
    records = []
    report = {
        "source_label": SOURCE_LABEL,
        "creator_count": len(creators),
        "print_contact_count": len(contacts),
        "inserted": 0,
        "merged": 0,
        "records": [],
    }

    for row in creators:
        existing = existing_by_name.get(slugify(row.name))
        record = build_creator_record(row, existing, geocode_cache)
        records.append(record)
        existing_by_name[slugify(row.name)] = record
        report[record["merge_action"]] += 1
        report["records"].append({
            "entity_name": record["entity_name"],
            "section": record["workbook_section"],
            "merge_action": record["merge_action"],
            "location_label": record["location_label"],
            "website_url": record["website_url"],
        })
        print(f"Creator: {record['entity_name']} -> {record['merge_action']}")

    for row in contacts:
        existing = existing_by_name.get(slugify(row.name))
        record = build_print_record(row, existing, geocode_cache)
        records.append(record)
        existing_by_name[slugify(row.name)] = record
        report[record["merge_action"]] += 1
        report["records"].append({
            "entity_name": record["entity_name"],
            "section": record["workbook_section"],
            "merge_action": record["merge_action"],
            "location_label": record["location_label"],
            "website_url": record["website_url"],
        })
        print(f"Print: {record['entity_name']} -> {record['merge_action']}")

    OUTPUT_SQL_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL_PATH.write_text(build_insert_sql(records) + "\n", encoding="utf-8")
    OUTPUT_REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote SQL migration to {OUTPUT_SQL_PATH}")
    print(f"Wrote import report to {OUTPUT_REPORT_PATH}")
    print(json.dumps({
        "creator_count": report["creator_count"],
        "print_contact_count": report["print_contact_count"],
        "inserted": report["inserted"],
        "merged": report["merged"],
        "total_records": len(records),
    }, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
