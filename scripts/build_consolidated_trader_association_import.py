import hashlib
import json
import re
import ssl
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\github\livelihood-ecosystem-directory")
WORKBOOK_T1 = Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\T1 - All_India_Trader_Associations_States_UTs.xlsx")
WORKBOOK_T2 = Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\T2 - india_trader_associations_100plus_single_table.xlsx")
WORKBOOK_T3 = Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\T3 - Comprehensive_Rural_Retail_Distribution_Networks_India.xlsx")
OUTPUT_SQL_PATH = PROJECT_ROOT / "supabase" / "migrations" / "20260514170000_import_consolidated_trader_associations.sql"
OUTPUT_REPORT_PATH = PROJECT_ROOT / "scripts" / "consolidated_trader_association_import_report.json"
CONFIG_PATH = PROJECT_ROOT / "config.js"
PUBLIC_VIEW = "ecosystem_directory_entities"
USER_AGENT = "Livelihood Ecosystem Directory Consolidated Trader Association Import/1.0"
SSL_CONTEXT = ssl._create_unverified_context()

SOURCE_LABEL = "Consolidated trader association workbooks"
CREATED_BY_NAME = "Consolidated trader association import"
CREATED_BY_EMAIL = "tanmay@greenruraleconomy.in"
INDIA_CENTER = (22.3511148, 78.6677428)

STATE_NAMES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh", "Goa", "Gujarat",
    "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka", "Kerala", "Madhya Pradesh",
    "Maharashtra", "Manipur", "Meghalaya", "Mizoram", "Nagaland", "Odisha", "Punjab",
    "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura", "Uttar Pradesh",
    "Uttarakhand", "West Bengal", "Delhi", "Jammu and Kashmir", "Ladakh", "Puducherry",
    "Chandigarh", "Andaman and Nicobar Islands", "Lakshadweep",
]


def compact_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_unicode_text(value: str) -> str:
    return (
        str(value or "")
        .replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2011", "-")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u00a0", " ")
        .replace("\u200b", "")
    )


def clean_text(value):
    text = compact_spaces(normalize_unicode_text(value))
    if not text or text.lower() in {"na", "n/a", "none", "null", "-", "--", "not available publicly"}:
        return None
    return text


def slugify(value: str) -> str:
    value = re.sub(r"[^\w\s-]", "", normalize_unicode_text(value or "").strip().lower())
    value = re.sub(r"[-\s]+", "-", value)
    return value.strip("-")


def dedupe(values):
    seen = set()
    output = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        output.append(text)
    return output


def split_list(value):
    text = clean_text(value)
    if not text:
        return []
    return dedupe(re.split(r"\s*[,;/|]\s*", text))


def sql_text(value):
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_text_array(values):
    cleaned = [str(item).replace("'", "''") for item in values if clean_text(item)]
    if not cleaned:
        return "'{}'::text[]"
    return "ARRAY[" + ", ".join(f"'{item}'" for item in cleaned) + "]::text[]"


def sql_json(value):
    return sql_text(json.dumps(value, ensure_ascii=False)) + "::jsonb"


def read_config():
    text = CONFIG_PATH.read_text(encoding="utf-8")
    url_match = re.search(r"SUPABASE_URL:\s*'([^']+)'", text)
    key_match = re.search(r"SUPABASE_ANON_KEY:\s*'([^']+)'", text)
    if not url_match or not key_match:
        raise RuntimeError("Could not read Supabase config from config.js")
    return url_match.group(1), key_match.group(1)


def fetch_existing_rows(base_url, anon_key):
    query = ",".join([
        "entity_uid",
        "entity_name",
        "summary",
        "description",
        "location_label",
        "primary_address",
        "district",
        "state",
        "country",
        "contact_email",
        "contact_phone",
        "website_url",
        "social_media",
        "office_locations",
        "tags",
        "keywords",
        "latitude",
        "longitude",
        "source_label",
        "source_url",
        "type_specific_data",
        "created_by_name",
        "created_by_email",
        "admin_notes",
        "search_text",
    ])
    url = (
        f"{base_url}/rest/v1/{PUBLIC_VIEW}"
        f"?select={urllib.parse.quote(query)}"
        f"&entity_type_slug=eq.trader_association&limit=5000"
    )
    headers = {
        "apikey": anon_key,
        "Authorization": f"Bearer {anon_key}",
        "Accept": "application/json",
    }
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
        return json.loads(response.read().decode("utf-8"))


def normalize_url(value):
    text = clean_text(value)
    if not text:
        return None
    if text.startswith("www."):
        text = "https://" + text
    elif re.match(r"^[A-Za-z0-9.-]+\.[A-Za-z]{2,}(/.*)?$", text):
        text = "https://" + text
    if not re.match(r"^https?://", text, flags=re.I):
        return None
    return text.rstrip("/")


def extract_email(value):
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"[\w.\-+%]+@[\w.\-]+\.[A-Za-z]{2,}", text, flags=re.I)
    return match.group(0) if match else None


def extract_phone(value):
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"(?:\+?\d[\d\s()./-]{7,}\d)", text)
    if not match:
        return None
    digits = re.sub(r"\D", "", match.group(0))
    if len(digits) < 10:
        return None
    return match.group(0).strip()


def merge_arrays(*arrays):
    merged = []
    for array in arrays:
        if isinstance(array, list):
            merged.extend(array)
    return dedupe(merged)


def pick_richer(primary, secondary):
    primary = clean_text(primary)
    secondary = clean_text(secondary)
    if not primary:
        return secondary
    if not secondary:
        return primary
    return secondary if len(secondary) > len(primary) else primary


def normalize_geography_tokens(*values):
    items = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        text = text.replace("(", ", ").replace(")", "")
        text = text.replace("&", ",")
        parts = re.split(r"\s*[,;/|]\s*", text)
        items.extend(part for part in parts if clean_text(part))
    normalized = []
    for item in items:
        fixed = item.replace("UP", "Uttar Pradesh").replace("MP", "Madhya Pradesh")
        if fixed == "All States":
            fixed = "Pan-India"
        normalized.append(fixed)
    return dedupe(normalized)


def infer_state_from_text(*values):
    haystack = ", ".join(clean_text(v) or "" for v in values).lower()
    for state in STATE_NAMES:
        if state.lower() in haystack:
            return state
    if "pan-india" in haystack or "all states" in haystack or "national" in haystack:
        return "India"
    return None


def primary_place_for_geocoding(place):
    text = clean_text(place)
    if not text:
        return None
    text = text.split("(", 1)[0].strip(" ,")
    if "National office:" in text:
        text = text.split(":", 1)[1].strip()
    return text or clean_text(place)


def geocode_query(query, cache):
    query = clean_text(query)
    if not query:
        return None
    if query in cache:
        return cache[query]
    url = f"https://nominatim.openstreetmap.org/search?format=jsonv2&limit=1&q={urllib.parse.quote(query)}"
    request = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception:
        cache[query] = None
        return None
    if not data:
        cache[query] = None
        return None
    match = data[0]
    try:
        point = (float(match["lat"]), float(match["lon"]))
    except (TypeError, KeyError, ValueError):
        cache[query] = None
        return None
    cache[query] = point
    return point


def geocode_location(place, address, state, geography_items, cache):
    primary_place = primary_place_for_geocoding(place)
    queries = dedupe([
        address,
        primary_place,
        place,
        f"{state}, India" if state and state != "India" else state,
        *geography_items[:3],
        "India",
    ])
    generic = {"india", "pan-india", "all states", "national"}
    for query in queries:
        normalized = (clean_text(query) or "").lower()
        if normalized in generic:
            continue
        point = geocode_query(query, cache)
        if point:
            lat, lng = point
            if 6 <= lat <= 38 and 68 <= lng <= 98:
                if normalized != "india":
                    time.sleep(0.8)
                return point
    return INDIA_CENTER


def build_search_text(record):
    values = [
        record.get("entity_name"),
        record.get("summary"),
        record.get("description"),
        record.get("location_label"),
        record.get("primary_address"),
        record.get("district"),
        record.get("state"),
        record.get("country"),
        " ".join(record.get("tags") or []),
        " ".join(record.get("keywords") or []),
    ]
    for value in (record.get("type_specific_data") or {}).values():
        if isinstance(value, list):
            values.append(" ".join(str(item) for item in value if clean_text(item)))
        else:
            values.append(clean_text(value) or "")
    return compact_spaces(" ".join(item for item in values if clean_text(item)))


def build_entity_uid(name):
    digest = hashlib.md5(name.encode("utf-8")).hexdigest()[:8]
    return f"trader_association-{slugify(name)}-{digest}"


def load_t1_rows():
    workbook = load_workbook(WORKBOOK_T1, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        name = clean_text(row[0])
        if not name or name.lower() == "association name":
            continue
        contact_raw = clean_text(row[4])
        rows.append({
            "name": name,
            "level": clean_text(row[1]),
            "state_or_region": clean_text(row[2]),
            "locality": None,
            "entity_type": "Trader association",
            "rural_relevance": None,
            "thematic_areas": split_list(row[5]),
            "commodities": split_list(row[6]),
            "geography_served": normalize_geography_tokens(row[2]),
            "website_url": normalize_url(row[3]),
            "contact_email": extract_email(contact_raw),
            "contact_phone": extract_phone(contact_raw),
            "address": clean_text(row[2]),
            "contact_person": None,
            "member_base": clean_text(row[7]),
            "market_linkages": clean_text(row[8]),
            "key_services": [],
            "registration_status": "Unknown",
            "source_url": normalize_url(row[3]),
            "source_notes": "Imported from T1 workbook",
            "source_file": WORKBOOK_T1.name,
        })
    return rows


def load_t2_rows():
    workbook = load_workbook(WORKBOOK_T2, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        name = clean_text(row[1] if len(row) > 1 else None)
        if not name or name.lower() == "entity_name":
            continue
        rows.append({
            "name": name,
            "level": clean_text(row[2]),
            "state_or_region": clean_text(row[3]),
            "locality": clean_text(row[4]),
            "entity_type": clean_text(row[5]),
            "rural_relevance": clean_text(row[6]),
            "thematic_areas": split_list(row[7]),
            "commodities": split_list(row[8]),
            "geography_served": normalize_geography_tokens(row[3], row[9]),
            "website_url": normalize_url(row[10]),
            "contact_email": extract_email(row[11]),
            "contact_phone": extract_phone(row[12]),
            "address": clean_text(row[13]),
            "contact_person": clean_text(row[14]),
            "member_base": clean_text(row[15]),
            "market_linkages": clean_text(row[16]),
            "key_services": split_list(row[17]),
            "registration_status": "Registered" if clean_text(row[18]) and "verified" in clean_text(row[18]).lower() else "Unknown",
            "source_url": normalize_url(row[20]),
            "source_notes": clean_text(row[21]),
            "source_file": WORKBOOK_T2.name,
        })
    return rows


def load_t3_rows():
    workbook = load_workbook(WORKBOOK_T3, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        name = clean_text(row[0])
        if not name or name.lower() == "organization name":
            continue
        contact_raw = clean_text(row[4])
        rows.append({
            "name": name,
            "level": clean_text(row[1]),
            "state_or_region": clean_text(row[2]),
            "locality": None,
            "entity_type": "Rural retail / distribution network",
            "rural_relevance": "High: rural retail and distribution",
            "thematic_areas": split_list(row[5]),
            "commodities": split_list(row[6]),
            "geography_served": normalize_geography_tokens(row[2]),
            "website_url": normalize_url(row[3]),
            "contact_email": extract_email(contact_raw),
            "contact_phone": extract_phone(contact_raw),
            "address": clean_text(row[2]),
            "contact_person": None,
            "member_base": clean_text(row[7]),
            "market_linkages": clean_text(row[8]),
            "key_services": [],
            "registration_status": "Unknown",
            "source_url": normalize_url(row[3]),
            "source_notes": "Imported from T3 workbook",
            "source_file": WORKBOOK_T3.name,
        })
    return rows


def merge_raw_rows(base, incoming):
    merged = dict(base)
    merged["level"] = pick_richer(merged.get("level"), incoming.get("level"))
    merged["state_or_region"] = pick_richer(merged.get("state_or_region"), incoming.get("state_or_region"))
    merged["locality"] = pick_richer(merged.get("locality"), incoming.get("locality"))
    merged["entity_type"] = pick_richer(merged.get("entity_type"), incoming.get("entity_type"))
    merged["rural_relevance"] = pick_richer(merged.get("rural_relevance"), incoming.get("rural_relevance"))
    merged["thematic_areas"] = merge_arrays(merged.get("thematic_areas"), incoming.get("thematic_areas"))
    merged["commodities"] = merge_arrays(merged.get("commodities"), incoming.get("commodities"))
    merged["geography_served"] = merge_arrays(merged.get("geography_served"), incoming.get("geography_served"))
    merged["website_url"] = pick_richer(merged.get("website_url"), incoming.get("website_url"))
    merged["contact_email"] = pick_richer(merged.get("contact_email"), incoming.get("contact_email"))
    merged["contact_phone"] = pick_richer(merged.get("contact_phone"), incoming.get("contact_phone"))
    merged["address"] = pick_richer(merged.get("address"), incoming.get("address"))
    merged["contact_person"] = pick_richer(merged.get("contact_person"), incoming.get("contact_person"))
    merged["member_base"] = pick_richer(merged.get("member_base"), incoming.get("member_base"))
    merged["market_linkages"] = pick_richer(merged.get("market_linkages"), incoming.get("market_linkages"))
    merged["key_services"] = merge_arrays(merged.get("key_services"), incoming.get("key_services"))
    merged["registration_status"] = pick_richer(merged.get("registration_status"), incoming.get("registration_status"))
    merged["source_url"] = pick_richer(merged.get("source_url"), incoming.get("source_url"))
    merged["source_notes"] = pick_richer(merged.get("source_notes"), incoming.get("source_notes"))
    merged["source_files"] = merge_arrays(merged.get("source_files") or [merged.get("source_file")], [incoming.get("source_file")])
    return merged


def choose_existing_match(raw, existing_rows):
    name_key = slugify(raw["name"])
    email = (clean_text(raw.get("contact_email")) or "").lower()
    website = raw.get("website_url") or ""
    website_domain = urllib.parse.urlparse(website).netloc.lower().removeprefix("www.") if website else ""
    for row in existing_rows:
        if slugify(row.get("entity_name")) == name_key:
            return row
    if email:
        for row in existing_rows:
            if (clean_text(row.get("contact_email")) or "").lower() == email:
                return row
    if website_domain:
        for row in existing_rows:
            current = row.get("website_url") or ""
            current_domain = urllib.parse.urlparse(current).netloc.lower().removeprefix("www.") if current else ""
            if current_domain and current_domain == website_domain:
                return row
    return None


def build_summary(raw):
    commodity_text = ", ".join(raw.get("commodities")[:3]) if raw.get("commodities") else "trade and distribution"
    geography_text = ", ".join(raw.get("geography_served")[:3]) if raw.get("geography_served") else "India"
    level_text = raw.get("level") or raw.get("entity_type") or "trade network"
    return f"{raw['name']} is a {level_text.lower()} organisation covering {commodity_text} across {geography_text}."


def build_description(raw):
    parts = []
    if raw.get("entity_type"):
        parts.append(raw["entity_type"])
    if raw.get("rural_relevance"):
        parts.append(raw["rural_relevance"])
    if raw.get("market_linkages"):
        parts.append("Market linkages: " + raw["market_linkages"])
    if raw.get("source_notes"):
        parts.append(raw["source_notes"])
    return ". ".join(parts) if parts else None


def merge_type_specific(existing_ts, incoming_ts):
    merged = dict(existing_ts or {})
    for key, value in incoming_ts.items():
        if isinstance(value, list):
            merged[key] = merge_arrays(existing_ts.get(key) if existing_ts else [], value)
        else:
            merged[key] = pick_richer(existing_ts.get(key) if existing_ts else None, value)
    return merged


def build_record(raw, existing, geocode_cache):
    state = infer_state_from_text(raw.get("state_or_region"), raw.get("address"), raw.get("locality"))
    location_label = raw.get("locality") or raw.get("state_or_region") or raw.get("address") or "India"
    lat, lng = geocode_location(raw.get("locality") or raw.get("state_or_region"), raw.get("address"), state, raw.get("geography_served") or [], geocode_cache)
    type_specific = {
        "commodities_or_sectors": raw.get("commodities") or [],
        "geography_served": raw.get("geography_served") or [],
        "member_base": raw.get("member_base"),
        "market_linkages": raw.get("market_linkages"),
        "key_services": raw.get("key_services") or [],
        "registration_status": raw.get("registration_status") or "Unknown",
    }
    summary = build_summary(raw)
    description = build_description(raw)
    entity_uid = existing.get("entity_uid") if existing else build_entity_uid(raw["name"])
    if existing:
        type_specific = merge_type_specific(existing.get("type_specific_data") or {}, type_specific)
        summary = pick_richer(existing.get("summary"), summary)
        description = pick_richer(existing.get("description"), description)
        raw["website_url"] = pick_richer(existing.get("website_url"), raw.get("website_url"))
        if clean_text(existing.get("location_label")) == clean_text(location_label):
            lat = existing.get("latitude") or lat
            lng = existing.get("longitude") or lng
        location_label = pick_richer(existing.get("location_label"), location_label)
        state = pick_richer(existing.get("state"), state)
    record = {
        "entity_uid": entity_uid,
        "entity_name": raw["name"],
        "summary": summary,
        "description": description,
        "location_label": location_label,
        "primary_address": raw.get("address") or location_label,
        "district": None,
        "state": state if state != "India" else None,
        "country": "India",
        "contact_email": raw.get("contact_email") or (existing.get("contact_email") if existing else None),
        "contact_phone": raw.get("contact_phone") or (existing.get("contact_phone") if existing else None),
        "website_url": raw.get("website_url"),
        "social_media": existing.get("social_media") if existing else {},
        "office_locations": existing.get("office_locations") if existing else [],
        "tags": merge_arrays(
            (existing.get("tags") if existing else []) or [],
            raw.get("thematic_areas") or [],
            raw.get("commodities") or [],
            raw.get("geography_served") or [],
            [raw.get("entity_type"), SOURCE_LABEL],
        ),
        "keywords": merge_arrays(
            (existing.get("keywords") if existing else []) or [],
            raw.get("thematic_areas") or [],
            raw.get("commodities") or [],
            raw.get("geography_served") or [],
            raw.get("key_services") or [],
            [raw["name"], raw.get("member_base"), raw.get("entity_type")],
        ),
        "latitude": lat,
        "longitude": lng,
        "source_label": SOURCE_LABEL,
        "source_url": raw.get("source_url") or raw.get("website_url"),
        "type_specific_data": type_specific,
        "created_by_name": existing.get("created_by_name") if existing else CREATED_BY_NAME,
        "created_by_email": existing.get("created_by_email") if existing else CREATED_BY_EMAIL,
        "admin_notes": f"Imported from consolidated trader association workbooks | Source files: {', '.join(raw.get('source_files') or [])}",
        "approval_status": "approved",
        "approved_at": "now()",
        "approved_by": "admin",
        "is_deleted": False,
        "updated_at": "now()",
        "search_text": None,
        "merge_action": "merged" if existing else "inserted",
    }
    record["search_text"] = build_search_text(record)
    return record


def build_values_sql(record):
    return "(" + ", ".join([
        sql_text(record["entity_uid"]),
        sql_text(record["entity_name"]),
        sql_text(record["summary"]),
        sql_text(record["description"]),
        sql_text(record["location_label"]),
        sql_text(record["primary_address"]),
        sql_text(record["district"]),
        sql_text(record["state"]),
        sql_text(record["country"]),
        sql_text(record["contact_email"]),
        sql_text(record["contact_phone"]),
        sql_text(record["website_url"]),
        sql_json(record["social_media"]),
        sql_json(record["office_locations"]),
        sql_text_array(record["tags"]),
        sql_text_array(record["keywords"]),
        str(record["latitude"]) if record["latitude"] is not None else "null",
        str(record["longitude"]) if record["longitude"] is not None else "null",
        sql_text(record["source_label"]),
        sql_text(record["source_url"]),
        sql_json(record["type_specific_data"]),
        sql_text(record["created_by_name"]),
        sql_text(record["created_by_email"]),
        sql_text(record["admin_notes"]),
        "'approved'",
        "now()",
        "'admin'",
        "false",
        sql_text(record["search_text"]),
        "now()",
    ]) + ")"


def build_insert_sql(records):
    return """insert into public.trader_association_entities (
  entity_uid,
  entity_name,
  summary,
  description,
  location_label,
  primary_address,
  district,
  state,
  country,
  contact_email,
  contact_phone,
  website_url,
  social_media,
  office_locations,
  tags,
  keywords,
  latitude,
  longitude,
  source_label,
  source_url,
  type_specific_data,
  created_by_name,
  created_by_email,
  admin_notes,
  approval_status,
  approved_at,
  approved_by,
  is_deleted,
  search_text,
  updated_at
) values
""" + ",\n".join(build_values_sql(record) for record in records) + """
on conflict (entity_uid) do update set
  entity_name = excluded.entity_name,
  summary = excluded.summary,
  description = excluded.description,
  location_label = excluded.location_label,
  primary_address = excluded.primary_address,
  district = excluded.district,
  state = excluded.state,
  country = excluded.country,
  contact_email = excluded.contact_email,
  contact_phone = excluded.contact_phone,
  website_url = excluded.website_url,
  social_media = excluded.social_media,
  office_locations = excluded.office_locations,
  tags = excluded.tags,
  keywords = excluded.keywords,
  latitude = excluded.latitude,
  longitude = excluded.longitude,
  source_label = excluded.source_label,
  source_url = excluded.source_url,
  type_specific_data = excluded.type_specific_data,
  created_by_name = excluded.created_by_name,
  created_by_email = excluded.created_by_email,
  admin_notes = excluded.admin_notes,
  approval_status = 'approved',
  approved_at = now(),
  approved_by = 'admin',
  is_deleted = false,
  search_text = excluded.search_text,
  updated_at = now();
"""


def main():
    base_url, anon_key = read_config()
    existing_rows = fetch_existing_rows(base_url, anon_key)
    raw_by_slug = {}
    raw_row_count = 0
    for row in load_t1_rows() + load_t2_rows() + load_t3_rows():
        raw_row_count += 1
        key = slugify(row["name"])
        if key in raw_by_slug:
            raw_by_slug[key] = merge_raw_rows(raw_by_slug[key], row)
        else:
            row["source_files"] = [row.get("source_file")]
            raw_by_slug[key] = row
    geocode_cache = {}
    records = []
    report = {
        "source_label": SOURCE_LABEL,
        "raw_source_rows": raw_row_count,
        "unique_workbook_entities": len(raw_by_slug),
        "existing_live_trader_association_count": len(existing_rows),
        "inserted": 0,
        "merged": 0,
        "with_website": 0,
        "with_phone": 0,
        "records": [],
    }
    for raw in raw_by_slug.values():
        existing = choose_existing_match(raw, existing_rows)
        record = build_record(raw, existing, geocode_cache)
        records.append(record)
        report[record["merge_action"]] += 1
        if clean_text(record.get("website_url")):
            report["with_website"] += 1
        if clean_text(record.get("contact_phone")):
            report["with_phone"] += 1
        report["records"].append({
            "entity_name": record["entity_name"],
            "merge_action": record["merge_action"],
            "location_label": record["location_label"],
            "website_url": record["website_url"],
            "source_files": raw.get("source_files") or [],
        })
        print(f"{record['entity_name']} -> {record['merge_action']}")

    OUTPUT_SQL_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL_PATH.write_text(build_insert_sql(records) + "\n", encoding="utf-8")
    OUTPUT_REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote SQL migration to {OUTPUT_SQL_PATH}")
    print(f"Wrote import report to {OUTPUT_REPORT_PATH}")
    print(json.dumps({
        "raw_source_rows": report["raw_source_rows"],
        "unique_workbook_entities": report["unique_workbook_entities"],
        "existing_live_trader_association_count": report["existing_live_trader_association_count"],
        "inserted": report["inserted"],
        "merged": report["merged"],
        "with_website": report["with_website"],
        "with_phone": report["with_phone"],
        "total_records": len(records),
    }, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
