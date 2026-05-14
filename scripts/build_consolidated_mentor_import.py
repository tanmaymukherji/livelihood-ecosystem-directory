import hashlib
import json
import re
import ssl
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\github\livelihood-ecosystem-directory")
WORKBOOK_PATHS = [
    Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\M1 - Mentors_Rural_Landscape_India_V2.xlsx"),
    Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\M2 - India_Rural_Mentors_1000_Individuals.xlsx"),
    Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\M3 - Startup_India_MAARG_Rural_Mentors_1000.xlsx"),
    Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\M4 - Final_Startup_India_and_Social_Entrepreneurship_Mentor_Database.xlsx"),
]
OUTPUT_SQL_PATH = PROJECT_ROOT / "supabase" / "migrations" / "20260514113000_import_consolidated_mentors.sql"
OUTPUT_REPORT_PATH = PROJECT_ROOT / "scripts" / "consolidated_mentor_import_report.json"
CONFIG_PATH = PROJECT_ROOT / "config.js"
PUBLIC_VIEW = "ecosystem_directory_entities"
USER_AGENT = "Livelihood Ecosystem Directory Consolidated Mentor Import/1.0"
SSL_CONTEXT = ssl._create_unverified_context()

SOURCE_LABEL = "Consolidated mentor workbooks"
CREATED_BY_NAME = "Consolidated mentor import"
CREATED_BY_EMAIL = "tanmay@greenruraleconomy.in"
INDIA_CENTER = (22.3511148, 78.6677428)

FREE_EMAIL_DOMAINS = {
    "gmail.com",
    "yahoo.com",
    "outlook.com",
    "hotmail.com",
    "rediffmail.com",
    "icloud.com",
    "ymail.com",
    "proton.me",
    "protonmail.com",
}

INVALID_PHONE_MARKERS = {
    "see linkedin",
    "available on request",
    "available on request (maarg)",
    "on request",
    "na",
    "n/a",
}

GENERIC_INDIA_LOCATIONS = {
    "india",
    "pan india",
    "pan-india",
    "rural india",
    "north india",
    "south india",
    "east india",
    "west india",
    "hybrid",
    "online",
    "remote",
}

STATE_NAMES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh", "Goa", "Gujarat",
    "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka", "Kerala", "Madhya Pradesh",
    "Maharashtra", "Manipur", "Meghalaya", "Mizoram", "Nagaland", "Odisha", "Punjab",
    "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura", "Uttar Pradesh",
    "Uttarakhand", "West Bengal", "Delhi", "Jammu and Kashmir", "Ladakh", "Puducherry",
    "Chandigarh", "Andaman and Nicobar Islands", "Lakshadweep",
]


def compact_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_unicode_text(value: str) -> str:
    return (
        str(value or "")
        .replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2011", "-")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u00a0", " ")
        .replace("\u200b", "")
    )


def clean_text(value):
    text = compact_spaces(normalize_unicode_text(value))
    if not text or text.lower() in {"na", "n/a", "none", "null", "-", "--"}:
        return None
    return text


def slugify(value: str) -> str:
    value = re.sub(r"[^\w\s-]", "", normalize_unicode_text(value or "").strip().lower())
    value = re.sub(r"[-\s]+", "-", value)
    return value.strip("-")


def dedupe(values):
    seen = set()
    output = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        output.append(text)
    return output


def split_list(value):
    text = clean_text(value)
    if not text:
        return []
    return dedupe(re.split(r"\s*[,;/|]\s*", text))


def split_rich_list(value):
    text = clean_text(value)
    if not text:
        return []
    text = text.replace("&", ",")
    return dedupe(re.split(r"\s*[,;/|]\s*", text))


def sql_text(value):
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_text_array(values):
    cleaned = [str(item).replace("'", "''") for item in values if clean_text(item)]
    if not cleaned:
        return "'{}'::text[]"
    return "ARRAY[" + ", ".join(f"'{item}'" for item in cleaned) + "]::text[]"


def sql_json(value):
    return sql_text(json.dumps(value, ensure_ascii=False)) + "::jsonb"


def read_config():
    text = CONFIG_PATH.read_text(encoding="utf-8")
    url_match = re.search(r"SUPABASE_URL:\s*'([^']+)'", text)
    key_match = re.search(r"SUPABASE_ANON_KEY:\s*'([^']+)'", text)
    if not url_match or not key_match:
        raise RuntimeError("Could not read Supabase config from config.js")
    return url_match.group(1), key_match.group(1)


def fetch_existing_mentors(base_url, anon_key):
    query = ",".join([
        "entity_uid",
        "entity_name",
        "summary",
        "description",
        "location_label",
        "primary_address",
        "district",
        "state",
        "country",
        "contact_email",
        "contact_phone",
        "website_url",
        "social_media",
        "office_locations",
        "tags",
        "keywords",
        "latitude",
        "longitude",
        "source_label",
        "source_url",
        "type_specific_data",
        "created_by_name",
        "created_by_email",
        "admin_notes",
        "search_text",
    ])
    url = (
        f"{base_url}/rest/v1/{PUBLIC_VIEW}"
        f"?select={urllib.parse.quote(query)}"
        f"&entity_type_slug=eq.mentor&limit=5000"
    )
    headers = {
        "apikey": anon_key,
        "Authorization": f"Bearer {anon_key}",
        "Accept": "application/json",
    }
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
        return json.loads(response.read().decode("utf-8"))


def normalize_url(value):
    text = clean_text(value)
    if not text:
        return None
    if " / " in text and "http" not in text and "." not in text.split(" / ")[0]:
        return None
    if text.startswith("www."):
        text = "https://" + text
    elif re.match(r"^[A-Za-z0-9.-]+\.[A-Za-z]{2,}(/.*)?$", text):
        text = "https://" + text
    elif text.startswith("linkedin.com") or text.startswith("twitter.com") or text.startswith("x.com") or text.startswith("facebook.com") or text.startswith("instagram.com") or text.startswith("youtube.com") or text.startswith("youtu.be"):
        text = "https://" + text
    if not re.match(r"^https?://", text, flags=re.I):
        return None
    return text.rstrip("/")


def parse_social_media(value):
    url = normalize_url(value)
    if not url:
        return {}
    lower = url.lower()
    if "linkedin.com" in lower:
        return {"linkedin": url}
    if "youtube.com" in lower or "youtu.be" in lower:
        return {"youtube": url}
    if "instagram.com" in lower:
        return {"instagram": url}
    if "facebook.com" in lower:
        return {"facebook": url}
    if "x.com" in lower or "twitter.com" in lower:
        return {"twitter": url}
    return {"other": [url]}


def merge_social_media(left, right):
    merged = {}
    for source in [left or {}, right or {}]:
        for key, value in source.items():
            if isinstance(value, list):
                merged[key] = dedupe((merged.get(key) or []) + value)
            else:
                merged[key] = value or merged.get(key)
    return merged


def merge_arrays(*arrays):
    merged = []
    for array in arrays:
        if isinstance(array, list):
            merged.extend(array)
    return dedupe(merged)


def pick_richer(primary, secondary):
    primary = clean_text(primary)
    secondary = clean_text(secondary)
    if not primary:
        return secondary
    if not secondary:
        return primary
    return secondary if len(secondary) > len(primary) else primary


def parse_phone(value):
    text = clean_text(value)
    if not text:
        return None
    lowered = text.lower()
    if any(marker in lowered for marker in INVALID_PHONE_MARKERS):
        return None
    digits = re.sub(r"\D", "", text)
    if len(digits) < 10:
        return None
    return text


def parse_years_experience(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"(\d{1,2})", text)
    return int(match.group(1)) if match else None


def professional_website_from_email(email):
    email = clean_text(email)
    if not email or "@" not in email:
        return None
    domain = email.split("@", 1)[1].lower()
    if domain in FREE_EMAIL_DOMAINS:
        return None
    return f"https://{domain}"


def fetch_homepage_title(url):
    if not url:
        return None, None
    # Keep bulk mentor imports fast and deterministic: we retain the website URL
    # for profile enrichment but skip per-record homepage fetches across hundreds
    # of domains.
    return url, None


def normalize_state_token(value):
    text = clean_text(value)
    if not text:
        return None
    normalized = text.replace("J&K", "Jammu and Kashmir").replace("UP", "Uttar Pradesh")
    for state in STATE_NAMES:
        if state.lower() == normalized.lower():
            return state
    return normalized


def normalize_geography_list(place, geography):
    items = []
    for value in [place, geography]:
        text = clean_text(value)
        if not text:
            continue
        parts = re.split(r"\s*[,;/|]\s*", text.replace(" + ", ", ").replace(" / ", ", "))
        items.extend(part for part in parts if clean_text(part))
    normalized = []
    for item in items:
        fixed = normalize_state_token(item)
        normalized.append(fixed)
    return dedupe(normalized)


def parse_state_and_location(place, geography_items):
    joined = ", ".join([clean_text(place) or "", ", ".join(geography_items)]).lower()
    state = None
    for candidate in STATE_NAMES:
        if candidate.lower() in joined:
            state = candidate
            break
    location_label = clean_text(place) or (", ".join(geography_items) if geography_items else "India")
    if not location_label:
        location_label = state or "India"
    return state, location_label


def geocode_query(query, cache):
    query = clean_text(query)
    if not query:
        return None
    if query in cache:
        return cache[query]
    url = f"https://nominatim.openstreetmap.org/search?format=jsonv2&limit=1&q={urllib.parse.quote(query)}"
    request = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception:
        cache[query] = None
        return None
    if not data:
        cache[query] = None
        return None
    match = data[0]
    try:
        point = (float(match["lat"]), float(match["lon"]))
    except (TypeError, KeyError, ValueError):
        cache[query] = None
        return None
    cache[query] = point
    return point


def geocode_location(place, location_label, state, cache):
    generic = {item.lower() for item in GENERIC_INDIA_LOCATIONS}
    queries = dedupe([
        place,
        location_label,
        f"{state}, India" if state else None,
        "India",
    ])
    for query in queries:
        normalized = (clean_text(query) or "").lower()
        if normalized in generic:
            continue
        point = geocode_query(query, cache)
        if point:
            lat, lng = point
            if 6 <= lat <= 38 and 68 <= lng <= 98:
                if normalized != "india":
                    time.sleep(0.8)
                return point
    return INDIA_CENTER


def build_search_text(record):
    values = [
        record.get("entity_name"),
        record.get("summary"),
        record.get("description"),
        record.get("location_label"),
        record.get("primary_address"),
        record.get("district"),
        record.get("state"),
        record.get("country"),
        " ".join(record.get("tags") or []),
        " ".join(record.get("keywords") or []),
    ]
    for value in (record.get("type_specific_data") or {}).values():
        if isinstance(value, list):
            values.append(" ".join(str(item) for item in value if clean_text(item)))
        else:
            values.append(clean_text(value) or "")
    return compact_spaces(" ".join(item for item in values if clean_text(item)))


def build_entity_uid(name):
    digest = hashlib.md5(name.encode("utf-8")).hexdigest()[:8]
    return f"mentor-{slugify(name)}-{digest}"


def load_workbook_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        name = clean_text(row[0])
        if not name or name.lower() == "name":
            continue
        if path.name.startswith("M4"):
            website_value = clean_text(row[3])
            rows.append({
                "name": name,
                "place": clean_text(row[1]),
                "email": clean_text(row[2]),
                "phone": None,
                "website_url": normalize_url(website_value),
                "organisation_hint": None if normalize_url(website_value) else website_value,
                "social_media": {},
                "domain_expertise": split_rich_list(row[4]),
                "industry_experience": [clean_text(row[5])] if clean_text(row[5]) else [],
                "years_experience": parse_years_experience(row[6]),
                "languages_spoken": split_rich_list(row[7]),
                "mentoring_modes": split_rich_list(row[8]),
                "geography_served": normalize_geography_list(row[1], row[9]),
                "target_stage": split_rich_list(row[10]),
                "availability_notes": clean_text(row[11]),
                "source_file": path.name,
            })
        else:
            rows.append({
                "name": name,
                "place": clean_text(row[1]),
                "email": clean_text(row[2]),
                "phone": parse_phone(row[3]),
                "website_url": normalize_url(row[4]),
                "organisation_hint": None,
                "social_media": parse_social_media(row[5]),
                "domain_expertise": split_rich_list(row[6]),
                "industry_experience": [clean_text(row[7])] if clean_text(row[7]) else [],
                "years_experience": parse_years_experience(row[8]),
                "languages_spoken": split_rich_list(row[9]),
                "mentoring_modes": split_rich_list(row[10]),
                "geography_served": normalize_geography_list(row[1], row[11]),
                "target_stage": split_rich_list(row[12]),
                "availability_notes": clean_text(row[13]),
                "source_file": path.name,
            })
    return rows


def merge_raw_rows(base, incoming):
    merged = dict(base)
    merged["place"] = pick_richer(merged.get("place"), incoming.get("place"))
    merged["email"] = pick_richer(merged.get("email"), incoming.get("email"))
    merged["phone"] = pick_richer(merged.get("phone"), incoming.get("phone"))
    merged["website_url"] = pick_richer(merged.get("website_url"), incoming.get("website_url"))
    merged["organisation_hint"] = pick_richer(merged.get("organisation_hint"), incoming.get("organisation_hint"))
    merged["social_media"] = merge_social_media(merged.get("social_media"), incoming.get("social_media"))
    merged["domain_expertise"] = merge_arrays(merged.get("domain_expertise"), incoming.get("domain_expertise"))
    merged["industry_experience"] = merge_arrays(merged.get("industry_experience"), incoming.get("industry_experience"))
    merged["years_experience"] = max(filter(None, [merged.get("years_experience"), incoming.get("years_experience")]), default=None)
    merged["languages_spoken"] = merge_arrays(merged.get("languages_spoken"), incoming.get("languages_spoken"))
    merged["mentoring_modes"] = merge_arrays(merged.get("mentoring_modes"), incoming.get("mentoring_modes"))
    merged["geography_served"] = merge_arrays(merged.get("geography_served"), incoming.get("geography_served"))
    merged["target_stage"] = merge_arrays(merged.get("target_stage"), incoming.get("target_stage"))
    merged["availability_notes"] = pick_richer(merged.get("availability_notes"), incoming.get("availability_notes"))
    merged["source_files"] = merge_arrays(merged.get("source_files") or [merged.get("source_file")], [incoming.get("source_file")])
    return merged


def choose_existing_match(raw, existing_rows):
    name_key = slugify(raw["name"])
    email = (clean_text(raw.get("email")) or "").lower()
    website = normalize_url(raw.get("website_url")) or professional_website_from_email(raw.get("email"))
    website_domain = urllib.parse.urlparse(website).netloc.lower().removeprefix("www.") if website else ""
    for row in existing_rows:
        if slugify(row.get("entity_name")) == name_key:
            return row
    if email:
        for row in existing_rows:
            if (clean_text(row.get("contact_email")) or "").lower() == email:
                return row
    if website_domain:
        for row in existing_rows:
            current = normalize_url(row.get("website_url"))
            current_domain = urllib.parse.urlparse(current).netloc.lower().removeprefix("www.") if current else ""
            if current_domain and current_domain == website_domain:
                return row
    return None


def build_summary(raw, organisation_title):
    focus = ", ".join(raw.get("domain_expertise")[:3]) if raw.get("domain_expertise") else "rural entrepreneurship"
    geographies = ", ".join(raw.get("geography_served")[:3]) if raw.get("geography_served") else "India"
    summary = f"{raw['name']} mentors on {focus} across {geographies}."
    if organisation_title:
        summary = f"{summary} Associated with {organisation_title}."
    return summary


def build_description(raw, organisation_title):
    parts = []
    if raw.get("organisation_hint"):
        parts.append(f"Associated organisation(s): {raw['organisation_hint']}")
    if organisation_title and organisation_title not in " ".join(parts):
        parts.append(f"Website title: {organisation_title}")
    if raw.get("industry_experience"):
        parts.append("Industry experience: " + ", ".join(raw["industry_experience"]))
    if raw.get("years_experience") is not None:
        parts.append(f"Years of experience: {raw['years_experience']}")
    if raw.get("availability_notes"):
        parts.append(f"Availability: {raw['availability_notes']}")
    return ". ".join(parts) if parts else None


def merge_type_specific(existing_ts, incoming_ts):
    merged = dict(existing_ts or {})
    for key, value in incoming_ts.items():
        if isinstance(value, list):
            merged[key] = merge_arrays(existing_ts.get(key) if existing_ts else [], value)
        else:
            merged[key] = pick_richer(existing_ts.get(key) if existing_ts else None, value)
    years = [existing_ts.get("years_experience") if existing_ts else None, incoming_ts.get("years_experience")]
    years = [item for item in years if item is not None]
    merged["years_experience"] = max(years) if years else None
    return merged


def build_record(raw, existing, geocode_cache):
    state, location_label = parse_state_and_location(raw.get("place"), raw.get("geography_served") or [])
    inferred_website = raw.get("website_url") or professional_website_from_email(raw.get("email"))
    final_website_url, homepage_title = fetch_homepage_title(inferred_website)
    organisation_title = raw.get("organisation_hint") or homepage_title
    summary = build_summary(raw, organisation_title)
    description = build_description(raw, homepage_title)
    lat, lng = geocode_location(raw.get("place"), location_label, state, geocode_cache)
    social_media = dict(raw.get("social_media") or {})
    if final_website_url and not social_media.get("website"):
        social_media["website"] = final_website_url
    type_specific = {
        "domain_expertise": raw.get("domain_expertise") or [],
        "industry_experience": raw.get("industry_experience") or [],
        "years_experience": raw.get("years_experience"),
        "languages_spoken": raw.get("languages_spoken") or [],
        "mentoring_modes": raw.get("mentoring_modes") or [],
        "geography_served": raw.get("geography_served") or [],
        "target_stage": raw.get("target_stage") or [],
        "availability_notes": raw.get("availability_notes"),
        "associated_organisations": split_rich_list(raw.get("organisation_hint")),
    }
    source_url = final_website_url or social_media.get("linkedin") or social_media.get("website")
    entity_uid = (existing or {}).get("entity_uid") or build_entity_uid(raw["name"])
    if existing:
        existing_ts = existing.get("type_specific_data") or {}
        type_specific = merge_type_specific(existing_ts, type_specific)
        social_media = merge_social_media(existing.get("social_media"), social_media)
        summary = pick_richer(existing.get("summary"), summary)
        description = pick_richer(existing.get("description"), description)
        final_website_url = pick_richer(existing.get("website_url"), final_website_url)
        if clean_text(existing.get("location_label")) == clean_text(location_label):
            lat = existing.get("latitude") or lat
            lng = existing.get("longitude") or lng
        location_label = pick_richer(existing.get("location_label"), location_label)
        state = pick_richer(existing.get("state"), state)
    record = {
        "entity_uid": entity_uid,
        "entity_name": raw["name"],
        "summary": summary,
        "description": description,
        "location_label": location_label,
        "primary_address": raw.get("place") or location_label,
        "district": None,
        "state": state,
        "country": "India",
        "contact_email": raw.get("email") or (existing.get("contact_email") if existing else None),
        "contact_phone": raw.get("phone") or (existing.get("contact_phone") if existing else None),
        "website_url": final_website_url,
        "social_media": social_media,
        "office_locations": existing.get("office_locations") if existing else [],
        "tags": merge_arrays(
            (existing.get("tags") if existing else []) or [],
            raw.get("domain_expertise") or [],
            raw.get("industry_experience") or [],
            raw.get("geography_served") or [],
            split_rich_list(raw.get("organisation_hint")),
            [SOURCE_LABEL],
        ),
        "keywords": merge_arrays(
            (existing.get("keywords") if existing else []) or [],
            raw.get("domain_expertise") or [],
            raw.get("industry_experience") or [],
            raw.get("geography_served") or [],
            raw.get("target_stage") or [],
            [raw["name"]],
        ),
        "latitude": lat,
        "longitude": lng,
        "source_label": SOURCE_LABEL,
        "source_url": source_url,
        "type_specific_data": type_specific,
        "created_by_name": existing.get("created_by_name") if existing else CREATED_BY_NAME,
        "created_by_email": existing.get("created_by_email") if existing else CREATED_BY_EMAIL,
        "admin_notes": f"Imported from consolidated mentor workbooks | Source files: {', '.join(raw.get('source_files') or [])}",
        "approval_status": "approved",
        "approved_at": "now()",
        "approved_by": "admin",
        "is_deleted": False,
        "updated_at": "now()",
        "search_text": None,
        "merge_action": "merged" if existing else "inserted",
    }
    record["search_text"] = build_search_text(record)
    return record


def build_values_sql(record):
    return "(" + ", ".join([
        sql_text(record["entity_uid"]),
        sql_text(record["entity_name"]),
        sql_text(record["summary"]),
        sql_text(record["description"]),
        sql_text(record["location_label"]),
        sql_text(record["primary_address"]),
        sql_text(record["district"]),
        sql_text(record["state"]),
        sql_text(record["country"]),
        sql_text(record["contact_email"]),
        sql_text(record["contact_phone"]),
        sql_text(record["website_url"]),
        sql_json(record["social_media"]),
        sql_json(record["office_locations"]),
        sql_text_array(record["tags"]),
        sql_text_array(record["keywords"]),
        str(record["latitude"]) if record["latitude"] is not None else "null",
        str(record["longitude"]) if record["longitude"] is not None else "null",
        sql_text(record["source_label"]),
        sql_text(record["source_url"]),
        sql_json(record["type_specific_data"]),
        sql_text(record["created_by_name"]),
        sql_text(record["created_by_email"]),
        sql_text(record["admin_notes"]),
        "'approved'",
        "now()",
        "'admin'",
        "false",
        sql_text(record["search_text"]),
        "now()",
    ]) + ")"


def build_insert_sql(records):
    return """insert into public.mentor_entities (
  entity_uid,
  entity_name,
  summary,
  description,
  location_label,
  primary_address,
  district,
  state,
  country,
  contact_email,
  contact_phone,
  website_url,
  social_media,
  office_locations,
  tags,
  keywords,
  latitude,
  longitude,
  source_label,
  source_url,
  type_specific_data,
  created_by_name,
  created_by_email,
  admin_notes,
  approval_status,
  approved_at,
  approved_by,
  is_deleted,
  search_text,
  updated_at
) values
""" + ",\n".join(build_values_sql(record) for record in records) + """
on conflict (entity_uid) do update set
  entity_name = excluded.entity_name,
  summary = excluded.summary,
  description = excluded.description,
  location_label = excluded.location_label,
  primary_address = excluded.primary_address,
  district = excluded.district,
  state = excluded.state,
  country = excluded.country,
  contact_email = excluded.contact_email,
  contact_phone = excluded.contact_phone,
  website_url = excluded.website_url,
  social_media = excluded.social_media,
  office_locations = excluded.office_locations,
  tags = excluded.tags,
  keywords = excluded.keywords,
  latitude = excluded.latitude,
  longitude = excluded.longitude,
  source_label = excluded.source_label,
  source_url = excluded.source_url,
  type_specific_data = excluded.type_specific_data,
  created_by_name = excluded.created_by_name,
  created_by_email = excluded.created_by_email,
  admin_notes = excluded.admin_notes,
  approval_status = 'approved',
  approved_at = now(),
  approved_by = 'admin',
  is_deleted = false,
  search_text = excluded.search_text,
  updated_at = now();
"""


def main():
    base_url, anon_key = read_config()
    existing_rows = fetch_existing_mentors(base_url, anon_key)
    raw_by_slug = {}
    raw_row_count = 0
    for path in WORKBOOK_PATHS:
        for row in load_workbook_rows(path):
            raw_row_count += 1
            key = slugify(row["name"])
            if key in raw_by_slug:
                raw_by_slug[key] = merge_raw_rows(raw_by_slug[key], row)
            else:
                row["source_files"] = [row.get("source_file")]
                raw_by_slug[key] = row
    geocode_cache = {}
    records = []
    report = {
        "source_label": SOURCE_LABEL,
        "raw_source_rows": raw_row_count,
        "unique_workbook_entities": len(raw_by_slug),
        "existing_live_mentor_count": len(existing_rows),
        "inserted": 0,
        "merged": 0,
        "with_website": 0,
        "with_social_media": 0,
        "with_phone": 0,
        "records": [],
    }
    for raw in raw_by_slug.values():
        existing = choose_existing_match(raw, existing_rows)
        record = build_record(raw, existing, geocode_cache)
        records.append(record)
        report[record["merge_action"]] += 1
        if clean_text(record.get("website_url")):
            report["with_website"] += 1
        if any(record.get("social_media", {}).values()):
            report["with_social_media"] += 1
        if clean_text(record.get("contact_phone")):
            report["with_phone"] += 1
        report["records"].append({
            "entity_name": record["entity_name"],
            "merge_action": record["merge_action"],
            "location_label": record["location_label"],
            "website_url": record["website_url"],
            "source_files": raw.get("source_files") or [],
        })
        print(f"{record['entity_name']} -> {record['merge_action']}")

    OUTPUT_SQL_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL_PATH.write_text(build_insert_sql(records) + "\n", encoding="utf-8")
    OUTPUT_REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote SQL migration to {OUTPUT_SQL_PATH}")
    print(f"Wrote import report to {OUTPUT_REPORT_PATH}")
    print(json.dumps({
        "raw_source_rows": report["raw_source_rows"],
        "unique_workbook_entities": report["unique_workbook_entities"],
        "existing_live_mentor_count": report["existing_live_mentor_count"],
        "inserted": report["inserted"],
        "merged": report["merged"],
        "with_website": report["with_website"],
        "with_social_media": report["with_social_media"],
        "with_phone": report["with_phone"],
        "total_records": len(records),
    }, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
