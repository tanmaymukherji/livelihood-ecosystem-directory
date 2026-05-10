import hashlib
import json
import re
import ssl
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\github\livelihood-ecosystem-directory")
CSR_LIST_PATH = Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\CSR List.xlsx")
FUNDER_DB_PATH = Path(r"C:\Users\tmukh\OneDrive\Desktop\Livelihood Network Database\india_csr_philanthropy_funder_database_v1.xlsx")
OUTPUT_SQL_PATH = PROJECT_ROOT / "supabase" / "migrations" / "20260510201500_import_csr_philanthropy_lists.sql"
OUTPUT_REPORT_PATH = PROJECT_ROOT / "scripts" / "csr_philanthropy_import_report.json"
CONFIG_PATH = PROJECT_ROOT / "config.js"
PUBLIC_VIEW = "ecosystem_directory_entities"
USER_AGENT = "Livelihood Ecosystem Directory CSR Import/1.0"
SSL_CONTEXT = ssl._create_unverified_context()

SOURCE_LABEL = "CSR and philanthropy workbooks"
CREATED_BY_NAME = "CSR philanthropy import"
CREATED_BY_EMAIL = "tanmay@greenruraleconomy.in"
INDIA_CENTER = (22.3511148, 78.6677428)

SUPPORTED_INSTRUMENTS = [
    "CSR grant",
    "Philanthropic grant",
    "Challenge fund",
    "Technical assistance",
    "Employee volunteering",
    "In-kind support",
    "Capacity building",
    "Blended finance / catalytic capital",
]

STATE_NAMES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh", "Goa", "Gujarat",
    "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka", "Kerala", "Madhya Pradesh",
    "Maharashtra", "Manipur", "Meghalaya", "Mizoram", "Nagaland", "Odisha", "Punjab",
    "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura", "Uttar Pradesh",
    "Uttarakhand", "West Bengal", "Delhi", "Jammu and Kashmir", "Ladakh", "Puducherry",
    "Chandigarh", "Andaman and Nicobar Islands", "Lakshadweep",
]


def compact_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_unicode_text(value: str) -> str:
    return (
        str(value or "")
        .replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2011", "-")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u00a0", " ")
        .replace("\u200b", "")
    )


def clean_text(value):
    text = compact_spaces(normalize_unicode_text(value))
    if not text or text.lower() in {"na", "n/a", "none", "null", "-", "not publicly stated"}:
        return None
    return text


def slugify(value: str) -> str:
    value = re.sub(r"[^\w\s-]", "", normalize_unicode_text(value or "").strip().lower())
    value = re.sub(r"[-\s]+", "-", value)
    return value.strip("-")


def dedupe(values):
    seen = set()
    output = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        output.append(text)
    return output


def sql_text(value):
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_text_array(values):
    cleaned = [str(item).replace("'", "''") for item in values if clean_text(item)]
    if not cleaned:
        return "'{}'::text[]"
    return "ARRAY[" + ", ".join(f"'{item}'" for item in cleaned) + "]::text[]"


def sql_json(value):
    return sql_text(json.dumps(value, ensure_ascii=False)) + "::jsonb"


def split_multi(value):
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"\s*[;,/|]\s*", text)
    return dedupe(parts)


def first_url(value):
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"https?://[^\s;,\]]+", text)
    return match.group(0).rstrip(".,;") if match else None


def all_urls(value):
    text = clean_text(value)
    if not text:
        return []
    return dedupe(match.rstrip(".,;") for match in re.findall(r"https?://[^\s;,\]]+", text))


def parse_contact_details(value):
    text = clean_text(value)
    if not text:
        return None, None
    emails = dedupe(re.findall(r"[\w.\-+%]+@[\w.\-]+\.[A-Za-z]{2,}", text, flags=re.I))
    phones = []
    for match in re.findall(r"(?:\+?\d[\d\s()./-]{7,}\d)", text):
        digits = re.sub(r"\D", "", match)
        if 10 <= len(digits) <= 14:
            phones.append(match.strip())
    return (emails[0] if emails else None), (dedupe(phones)[0] if phones else None)


def merge_arrays(*arrays):
    merged = []
    for array in arrays:
        if isinstance(array, list):
            merged.extend(array)
    return dedupe(merged)


def pick_richer(primary, secondary):
    primary = clean_text(primary)
    secondary = clean_text(secondary)
    if not primary:
        return secondary
    if not secondary:
        return primary
    return secondary if len(secondary) > len(primary) else primary


def normalize_geographies(value):
    items = split_multi(value)
    normalized = []
    for item in items:
        fixed = (
            item.replace("Pan India", "Pan-India")
            .replace("pan India", "Pan-India")
            .replace("Pan India", "Pan-India")
        )
        normalized.append(fixed)
    return dedupe(normalized)


def map_support_instruments(value):
    text = clean_text(value)
    if not text:
        return []
    haystack = text.lower()
    mapped = []
    if "csr" in haystack:
        mapped.append("CSR grant")
    if "grant" in haystack or "philanth" in haystack:
        mapped.append("Philanthropic grant")
    if "challenge" in haystack:
        mapped.append("Challenge fund")
    if "technical" in haystack:
        mapped.append("Technical assistance")
    if "volunteer" in haystack:
        mapped.append("Employee volunteering")
    if "in-kind" in haystack or "in kind" in haystack:
        mapped.append("In-kind support")
    if "capacity" in haystack:
        mapped.append("Capacity building")
    if "catalytic" in haystack or "blended" in haystack:
        mapped.append("Blended finance / catalytic capital")
    return dedupe(mapped)


def infer_location_parts(address):
    text = clean_text(address)
    if not text:
        return None, None, None
    state = None
    lowered = text.lower()
    for name in sorted(STATE_NAMES, key=len, reverse=True):
        if name.lower() in lowered:
            state = name
            break
    district = None
    parts = [compact_spaces(part) for part in re.split(r"[;,/]", text) if compact_spaces(part)]
    if state:
        for index, part in enumerate(parts):
            if state.lower() in part.lower():
                if index > 0:
                    district = re.sub(r"\b\d{4,6}\b", "", parts[index - 1]).strip(" -") or None
                break
    location_label = ", ".join(part for part in [district, state] if clean_text(part)) or text
    return district, state, location_label


def geocode_query(query, cache):
    query = clean_text(query)
    if not query:
        return None
    if query in cache:
        return cache[query]
    url = f"https://nominatim.openstreetmap.org/search?format=jsonv2&limit=1&q={urllib.parse.quote(query)}"
    request = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception:
        cache[query] = None
        return None
    if not data:
        cache[query] = None
        return None
    match = data[0]
    try:
        point = (float(match["lat"]), float(match["lon"]))
    except (TypeError, KeyError, ValueError):
        cache[query] = None
        return None
    cache[query] = point
    return point


def geocode_location(address, state, geographies, cache):
    generic = {"india", "pan-india", "pan india", "online platform", "online"}
    queries = dedupe([
        address,
        f"{state}, India" if state else None,
        *geographies,
        "India",
    ])
    for query in queries:
        normalized = (clean_text(query) or "").lower()
        if normalized in generic:
            continue
        point = geocode_query(query, cache)
        if point:
            lat, lng = point
            if 6 <= lat <= 38 and 68 <= lng <= 98:
                if normalized != "india":
                    time.sleep(0.8)
                return point
    return INDIA_CENTER


def read_config():
    text = CONFIG_PATH.read_text(encoding="utf-8")
    url_match = re.search(r"SUPABASE_URL:\s*'([^']+)'", text)
    key_match = re.search(r"SUPABASE_ANON_KEY:\s*'([^']+)'", text)
    if not url_match or not key_match:
        raise RuntimeError("Could not read Supabase config from config.js")
    return url_match.group(1), key_match.group(1)


def fetch_existing_rows(base_url, anon_key):
    query = ",".join([
        "entity_uid",
        "entity_name",
        "summary",
        "description",
        "location_label",
        "primary_address",
        "district",
        "state",
        "country",
        "contact_email",
        "contact_phone",
        "website_url",
        "social_media",
        "office_locations",
        "tags",
        "keywords",
        "latitude",
        "longitude",
        "source_label",
        "source_url",
        "type_specific_data",
        "created_by_name",
        "created_by_email",
        "admin_notes",
        "search_text",
    ])
    url = (
        f"{base_url}/rest/v1/{PUBLIC_VIEW}"
        f"?select={urllib.parse.quote(query)}"
        f"&entity_type_slug=eq.csr_philanthropy&limit=1000"
    )
    headers = {
        "apikey": anon_key,
        "Authorization": f"Bearer {anon_key}",
        "Accept": "application/json",
    }
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
        return json.loads(response.read().decode("utf-8"))


def load_csr_list_rows():
    wb = load_workbook(CSR_LIST_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        name = clean_text(row[0])
        if not name:
            continue
        contact_email, contact_phone = parse_contact_details(row[2])
        rows.append({
            "name": name,
            "category": "CSR / Philanthropy",
            "address": clean_text(row[1]),
            "contact_details": clean_text(row[2]),
            "contact_email": contact_email,
            "contact_phone": contact_phone,
            "website_url": first_url(row[3]),
            "focus_areas": split_multi(row[4]),
            "geographies_served": normalize_geographies(row[5]),
            "grant_size": clean_text(row[6]),
            "support_instruments": map_support_instruments(row[7]),
            "partner_focus": split_multi(row[8]),
            "application_link": first_url(row[9]),
            "partnership_preference": clean_text(row[10]),
            "reporting_notes": clean_text(row[11]),
            "source_urls": dedupe([first_url(row[3]), first_url(row[9])]),
            "data_confidence": None,
            "last_checked": None,
            "source_sheet": "CSR List",
        })
    return rows


def load_funder_db_rows():
    wb = load_workbook(FUNDER_DB_PATH, read_only=True, data_only=True)
    ws = wb["Funder_Database"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        name = clean_text(row[1])
        if not name or name.lower().startswith("organisation / funder"):
            continue
        contact_email, contact_phone = parse_contact_details(row[4])
        rows.append({
            "name": name,
            "category": clean_text(row[2]),
            "address": clean_text(row[3]),
            "contact_details": clean_text(row[4]),
            "contact_email": contact_email,
            "contact_phone": contact_phone,
            "website_url": first_url(row[5]),
            "focus_areas": split_multi(row[6]),
            "geographies_served": normalize_geographies(row[7]),
            "grant_size": clean_text(row[8]),
            "support_instruments": map_support_instruments(row[9]),
            "partner_focus": split_multi(row[10]),
            "application_link": first_url(row[11]),
            "partnership_preference": clean_text(row[12]),
            "reporting_notes": clean_text(row[13]),
            "source_urls": all_urls(row[14]),
            "data_confidence": clean_text(row[15]),
            "last_checked": clean_text(row[16]),
            "source_sheet": "Funder_Database",
        })
    return rows


def merge_raw_rows(base_rows, incoming):
    merged = dict(base_rows)
    merged["category"] = pick_richer(merged.get("category"), incoming.get("category"))
    merged["address"] = pick_richer(merged.get("address"), incoming.get("address"))
    merged["contact_details"] = pick_richer(merged.get("contact_details"), incoming.get("contact_details"))
    merged["contact_email"] = pick_richer(merged.get("contact_email"), incoming.get("contact_email"))
    merged["contact_phone"] = pick_richer(merged.get("contact_phone"), incoming.get("contact_phone"))
    merged["website_url"] = pick_richer(merged.get("website_url"), incoming.get("website_url"))
    merged["focus_areas"] = merge_arrays(merged.get("focus_areas"), incoming.get("focus_areas"))
    merged["geographies_served"] = merge_arrays(merged.get("geographies_served"), incoming.get("geographies_served"))
    merged["grant_size"] = pick_richer(merged.get("grant_size"), incoming.get("grant_size"))
    merged["support_instruments"] = merge_arrays(merged.get("support_instruments"), incoming.get("support_instruments"))
    merged["partner_focus"] = merge_arrays(merged.get("partner_focus"), incoming.get("partner_focus"))
    merged["application_link"] = pick_richer(merged.get("application_link"), incoming.get("application_link"))
    merged["partnership_preference"] = pick_richer(merged.get("partnership_preference"), incoming.get("partnership_preference"))
    merged["reporting_notes"] = pick_richer(merged.get("reporting_notes"), incoming.get("reporting_notes"))
    merged["source_urls"] = merge_arrays(merged.get("source_urls"), incoming.get("source_urls"))
    merged["data_confidence"] = pick_richer(merged.get("data_confidence"), incoming.get("data_confidence"))
    merged["last_checked"] = pick_richer(merged.get("last_checked"), incoming.get("last_checked"))
    merged["source_sheet"] = merge_arrays([merged.get("source_sheet")], [incoming.get("source_sheet")])
    return merged


def choose_existing_match(raw, existing_rows):
    name_key = slugify(raw["name"])
    website = raw.get("website_url") or ""
    website_domain = urllib.parse.urlparse(website).netloc.lower().removeprefix("www.") if website else ""
    for row in existing_rows:
        if slugify(row.get("entity_name")) == name_key:
            return row
    if website_domain:
        for row in existing_rows:
            current = row.get("website_url") or ""
            current_domain = urllib.parse.urlparse(current).netloc.lower().removeprefix("www.") if current else ""
            if current_domain and current_domain == website_domain:
                return row
    return None


def build_summary(name, category, focus_areas, geographies):
    category_text = category or "CSR / philanthropy organisation"
    focus_text = ", ".join(focus_areas[:4]) if focus_areas else "multi-sector support"
    geo_text = ", ".join(geographies[:3]) if geographies else "India"
    return f"{name} is a {category_text.lower()} supporting {focus_text} across {geo_text}."


def build_description(raw):
    parts = [
        raw.get("category"),
        raw.get("grant_size"),
        clean_text("Support: " + ", ".join(raw.get("support_instruments") or [])) if raw.get("support_instruments") else None,
        clean_text("Partner focus: " + ", ".join(raw.get("partner_focus") or [])) if raw.get("partner_focus") else None,
        clean_text("Partnership preference: " + (raw.get("partnership_preference") or "")),
        clean_text("Compliance: " + (raw.get("reporting_notes") or "")),
    ]
    return ". ".join(part for part in parts if clean_text(part))


def build_search_text(record):
    values = [
        record.get("entity_name"),
        record.get("summary"),
        record.get("description"),
        record.get("location_label"),
        record.get("primary_address"),
        record.get("district"),
        record.get("state"),
        record.get("country"),
        " ".join(record.get("tags") or []),
        " ".join(record.get("keywords") or []),
    ]
    for value in (record.get("type_specific_data") or {}).values():
        if isinstance(value, list):
            values.append(" ".join(str(item) for item in value if clean_text(item)))
        else:
            values.append(clean_text(value) or "")
    return compact_spaces(" ".join(item for item in values if clean_text(item)))


def build_entity_uid(name):
    digest = hashlib.md5(name.encode("utf-8")).hexdigest()[:8]
    return f"csr_philanthropy-{slugify(name)}-{digest}"


def build_record(raw, existing, geocode_cache):
    district, state, inferred_location = infer_location_parts(raw.get("address"))
    location_label = inferred_location or (", ".join(raw.get("geographies_served") or []) or "India")
    lat, lng = geocode_location(raw.get("address"), state, raw.get("geographies_served") or [], geocode_cache)
    type_specific_data = {
        "focus_areas": raw.get("focus_areas") or [],
        "geography_served": raw.get("geographies_served") or [],
        "support_instruments": raw.get("support_instruments") or [],
        "typical_support_size": raw.get("grant_size"),
        "beneficiary_or_partner_focus": raw.get("partner_focus") or [],
        "application_or_nomination_process": raw.get("partnership_preference"),
        "application_link": raw.get("application_link"),
        "partnership_preferences": split_multi(raw.get("partnership_preference")),
        "reporting_or_compliance_notes": raw.get("reporting_notes"),
    }
    summary = build_summary(raw["name"], raw.get("category"), raw.get("focus_areas") or [], raw.get("geographies_served") or [])
    description = build_description(raw)
    entity_uid = existing.get("entity_uid") if existing else build_entity_uid(raw["name"])
    website_url = raw.get("website_url") or (existing.get("website_url") if existing else None)
    source_url = raw.get("application_link") or raw.get("website_url") or ((raw.get("source_urls") or [None])[0])
    if existing:
        existing_ts = existing.get("type_specific_data") or {}
        merged_ts = dict(existing_ts)
        for key, value in type_specific_data.items():
            if isinstance(value, list):
                merged_ts[key] = merge_arrays(existing_ts.get(key) or [], value)
            else:
                merged_ts[key] = pick_richer(existing_ts.get(key), value)
        type_specific_data = merged_ts
        summary = pick_richer(existing.get("summary"), summary)
        description = pick_richer(existing.get("description"), description)
        website_url = pick_richer(existing.get("website_url"), website_url)
        lat = existing.get("latitude") or lat
        lng = existing.get("longitude") or lng
    record = {
        "entity_uid": entity_uid,
        "entity_name": raw["name"],
        "summary": summary,
        "description": description,
        "location_label": location_label,
        "primary_address": raw.get("address"),
        "district": district,
        "state": state,
        "country": "India",
        "contact_email": raw.get("contact_email"),
        "contact_phone": raw.get("contact_phone"),
        "website_url": website_url,
        "social_media": existing.get("social_media") if existing else {},
        "office_locations": existing.get("office_locations") if existing else [],
        "tags": merge_arrays(
            (existing.get("tags") if existing else []) or [],
            raw.get("focus_areas") or [],
            raw.get("geographies_served") or [],
            [raw.get("category"), SOURCE_LABEL],
        ),
        "keywords": merge_arrays(
            (existing.get("keywords") if existing else []) or [],
            raw.get("focus_areas") or [],
            raw.get("partner_focus") or [],
            raw.get("geographies_served") or [],
            [raw["name"], raw.get("category")],
        ),
        "latitude": lat,
        "longitude": lng,
        "source_label": SOURCE_LABEL,
        "source_url": source_url,
        "type_specific_data": type_specific_data,
        "created_by_name": existing.get("created_by_name") if existing else CREATED_BY_NAME,
        "created_by_email": existing.get("created_by_email") if existing else CREATED_BY_EMAIL,
        "admin_notes": f"Imported from CSR List.xlsx and india_csr_philanthropy_funder_database_v1.xlsx | Source sheets: {', '.join(raw.get('source_sheet') or [])}",
        "approval_status": "approved",
        "approved_at": "now()",
        "approved_by": "admin",
        "is_deleted": False,
        "updated_at": "now()",
        "search_text": None,
        "merge_action": "merged" if existing else "inserted",
    }
    record["search_text"] = build_search_text(record)
    return record


def build_values_sql(record):
    return "(" + ", ".join([
        sql_text(record["entity_uid"]),
        sql_text(record["entity_name"]),
        sql_text(record["summary"]),
        sql_text(record["description"]),
        sql_text(record["location_label"]),
        sql_text(record["primary_address"]),
        sql_text(record["district"]),
        sql_text(record["state"]),
        sql_text(record["country"]),
        sql_text(record["contact_email"]),
        sql_text(record["contact_phone"]),
        sql_text(record["website_url"]),
        sql_json(record["social_media"]),
        sql_json(record["office_locations"]),
        sql_text_array(record["tags"]),
        sql_text_array(record["keywords"]),
        str(record["latitude"]) if record["latitude"] is not None else "null",
        str(record["longitude"]) if record["longitude"] is not None else "null",
        sql_text(record["source_label"]),
        sql_text(record["source_url"]),
        sql_json(record["type_specific_data"]),
        sql_text(record["created_by_name"]),
        sql_text(record["created_by_email"]),
        sql_text(record["admin_notes"]),
        "'approved'",
        "now()",
        "'admin'",
        "false",
        sql_text(record["search_text"]),
        "now()",
    ]) + ")"


def build_insert_sql(records):
    return """insert into public.csr_philanthropy_entities (
  entity_uid,
  entity_name,
  summary,
  description,
  location_label,
  primary_address,
  district,
  state,
  country,
  contact_email,
  contact_phone,
  website_url,
  social_media,
  office_locations,
  tags,
  keywords,
  latitude,
  longitude,
  source_label,
  source_url,
  type_specific_data,
  created_by_name,
  created_by_email,
  admin_notes,
  approval_status,
  approved_at,
  approved_by,
  is_deleted,
  search_text,
  updated_at
) values
""" + ",\n".join(build_values_sql(record) for record in records) + """
on conflict (entity_uid) do update set
  entity_name = excluded.entity_name,
  summary = excluded.summary,
  description = excluded.description,
  location_label = excluded.location_label,
  primary_address = excluded.primary_address,
  district = excluded.district,
  state = excluded.state,
  country = excluded.country,
  contact_email = excluded.contact_email,
  contact_phone = excluded.contact_phone,
  website_url = excluded.website_url,
  social_media = excluded.social_media,
  office_locations = excluded.office_locations,
  tags = excluded.tags,
  keywords = excluded.keywords,
  latitude = excluded.latitude,
  longitude = excluded.longitude,
  source_label = excluded.source_label,
  source_url = excluded.source_url,
  type_specific_data = excluded.type_specific_data,
  created_by_name = excluded.created_by_name,
  created_by_email = excluded.created_by_email,
  admin_notes = excluded.admin_notes,
  approval_status = 'approved',
  approved_at = now(),
  approved_by = 'admin',
  is_deleted = false,
  search_text = excluded.search_text,
  updated_at = now();
"""


def main():
    base_url, anon_key = read_config()
    existing_rows = fetch_existing_rows(base_url, anon_key)
    raw_rows = {}
    for source_row in load_csr_list_rows() + load_funder_db_rows():
        key = slugify(source_row["name"])
        if key in raw_rows:
            raw_rows[key] = merge_raw_rows(raw_rows[key], source_row)
        else:
            source_row["source_sheet"] = [source_row.get("source_sheet")]
            raw_rows[key] = source_row
    geocode_cache = {}
    records = []
    report = {
        "source_label": SOURCE_LABEL,
        "raw_source_rows": len(load_csr_list_rows()) + len(load_funder_db_rows()),
        "unique_workbook_entities": len(raw_rows),
        "inserted": 0,
        "merged": 0,
        "records": [],
    }
    for raw in raw_rows.values():
        existing = choose_existing_match(raw, existing_rows)
        record = build_record(raw, existing, geocode_cache)
        records.append(record)
        report[record["merge_action"]] += 1
        report["records"].append({
            "entity_name": record["entity_name"],
            "merge_action": record["merge_action"],
            "website_url": record["website_url"],
            "location_label": record["location_label"],
        })
        print(f"{record['entity_name']} -> {record['merge_action']}")

    OUTPUT_SQL_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL_PATH.write_text(build_insert_sql(records) + "\n", encoding="utf-8")
    OUTPUT_REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote SQL migration to {OUTPUT_SQL_PATH}")
    print(f"Wrote import report to {OUTPUT_REPORT_PATH}")
    print(json.dumps({
        "raw_source_rows": report["raw_source_rows"],
        "unique_workbook_entities": report["unique_workbook_entities"],
        "inserted": report["inserted"],
        "merged": report["merged"],
        "total_records": len(records),
    }, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
